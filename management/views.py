import io
import json
import re
import traceback
import unicodedata
from datetime import date, datetime, timedelta, timezone as dt_timezone
from decimal import Decimal, InvalidOperation
from xml.sax.saxutils import escape
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from invoices.models import Facture
from technique.models import TechnicalProject
from user_access.user_test_functions import has_administratif_access

from .email_manager import (
    create_outlook_event,
    delete_outlook_event,
    fetch_new_emails,
    get_email_summary,
    get_sent_emails,
    send_email_reply,
    update_outlook_event,
)
from .models import (
    Activite,
    CategorieDossierAdministratif,
    ChampPersonnaliseDossier,
    NotificationInterne,
    RegleRappelActivite,
    TypeActivite,
    GmailConversation,
    GmailConversationEvent,
    ValeurChampPersonnaliseDossier,
)
from .gmail_service import (
    send_conversation_reminder,
    sync_conversation_journal,
)

Utilisateur = get_user_model()

STATUS_COLORS = {
    "todo": "#64748b",
    "in_progress": "#2563eb",
    "done": "#16a34a",
    "cancelled": "#dc2626",
}

PRIORITY_COLORS = {
    "low": "#64748b",
    "normal": "#0ea5e9",
    "high": "#f97316",
    "urgent": "#dc2626",
}

ADMIN_PROJECT_EXPORT_COLUMNS = [
    ("Référence", "reference"),
    ("Affaire", "affaire"),
    ("Lot / étage", "lot_etage"),
    ("Adresse du bien", "adresse_bien"),
    ("Vendeur", "vendeur"),
    ("Bénéficiaire", "beneficiaire"),
    ("Locataire", "locataire"),
    ("Type de dossier", "type_dossier_label"),
    ("Activité métier", "activite_metier_label"),
    ("État", "etat_label"),
    ("Catégorie", "categorie_label"),
    ("Date de promesse", "date_promesse"),
    ("1ère période", "premiere_periode"),
    ("2ème période", "deuxieme_periode"),
    ("Négociation externe", "negociation_externe"),
    ("Frais", "frais"),
    ("Prix", "prix"),
    ("DG", "dg"),
    ("Date DG", "date_dg"),
    ("CS prêt", "cs_pret"),
    ("Date CS prêt", "date_cs_pret"),
    ("Date de réitération", "date_reiteration"),
    ("Acte", "acte"),
    ("Parcelles", "parcelles"),
    ("Dépôt permis", "depot_permis"),
    ("Obtention permis", "obtention_permis"),
    ("Diags", "diags"),
    ("Bornage", "bornage"),
    ("Étude sol / géotechnique", "etude_sol_geotechnique"),
    ("Étude pollution", "etude_pollution"),
    ("Étude d’impact", "etude_impact"),
    ("Prorogation", "prorogation"),
    ("Avenant 1", "avenant_1"),
    ("Avenant 2", "avenant_2"),
    ("Avenant 3", "avenant_3"),
    ("Relevés de compte", "releves_compte"),
]

ADMIN_PROJECT_IMPORT_ALIASES = {
    "reference": "reference",
    "ref": "reference",
    "affaire": "affaire",
    "affaires": "affaire",
    "nom du dossier": "affaire",
    "dossier": "affaire",
    "lots etage": "lot_etage",
    "lot etage": "lot_etage",
    "lot / etage": "lot_etage",
    "adresse du bien": "adresse_bien",
    "adresse": "adresse_bien",
    "vendeur": "vendeur",
    "beneficiaire": "beneficiaire",
    "locataire": "locataire",
    "type": "type_dossier",
    "type de dossier": "type_dossier",
    "activite": "activite_metier",
    "activite metier": "activite_metier",
    "etat": "etat",
    "categorie": "categorie",
    "date promesse": "date_promesse",
    "date de promesse": "date_promesse",
    "1ere periode": "premiere_periode",
    "1re periode": "premiere_periode",
    "premiere periode": "premiere_periode",
    "2eme periode": "deuxieme_periode",
    "deuxieme periode": "deuxieme_periode",
    "negociation externe": "negociation_externe",
    "frais": "frais",
    "prix": "prix",
    "dg": "dg",
    "depot de garantie": "dg",
    "date dg": "date_dg",
    "cs pret": "cs_pret",
    "conditions suspensives": "cs_pret",
    "conditions suspensives pret": "cs_pret",
    "date cs pret": "date_cs_pret",
    "date reiteration": "date_reiteration",
    "date de reiteration": "date_reiteration",
    "acte": "acte",
    "parcelles": "parcelles",
    "depot permis": "depot_permis",
    "obtention permis": "obtention_permis",
    "diags": "diags",
    "bornage": "bornage",
    "etude sol geotechnique": "etude_sol_geotechnique",
    "etude pollution": "etude_pollution",
    "etude impact": "etude_impact",
    "prorogation": "prorogation",
    "avenant 1": "avenant_1",
    "avenant 2": "avenant_2",
    "avenant 3": "avenant_3",
    "releves compte": "releves_compte",
    "releves de compte": "releves_compte",
}

ADMIN_PROJECT_PROMOTION_SHEET = "Promotion immobilière"
ADMIN_PROJECT_OTHER_SHEET = "Autres dossiers"


def _json_error(message, status=400):
    return JsonResponse({"success": False, "message": message}, status=status)


def _parse_request_json(request):
    try:
        return json.loads(request.body or "{}")
    except json.JSONDecodeError:
        raise ValueError("JSON invalide")


def _parse_iso_datetime(value):
    value = (value or "").strip()
    if not value:
        raise ValueError("Date manquante")

    if value.endswith("Z"):
        value = value[:-1] + "+00:00"

    parsed = datetime.fromisoformat(value)
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _parse_iso_date(value, label):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError(f"{label} doit être une date valide.")


def _parse_non_negative_decimal(value, label):
    raw = str(value or "0").strip()
    raw = raw.replace("\xa0", "").replace(" ", "").replace("€", "")
    if "," in raw and "." in raw and raw.rfind(",") > raw.rfind("."):
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    try:
        amount = Decimal(raw)
    except (InvalidOperation, ValueError):
        raise ValueError(f"{label} doit être un montant valide.")
    if amount < 0:
        raise ValueError(f"{label} ne peut pas être négatif.")
    return amount


def _truthy(value):
    if isinstance(value, bool):
        return value
    return str(value or "").lower() in {"1", "true", "on", "yes", "oui"}


def _normalize_header(value):
    raw = str(value or "").strip().lower()
    raw = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")
    for char in ["/", "-", "_", ".", ":", ";", "(", ")", "[", "]", "'"]:
        raw = raw.replace(char, " ")
    return " ".join(raw.split())


def _choice_value(value, choices, label):
    raw = str(value or "").strip()
    if not raw:
        return ""
    normalized = _normalize_header(raw)
    for key, display in choices:
        if normalized in {_normalize_header(key), _normalize_header(display)}:
            return key
    raise ValueError(f"{label} invalide : {raw}")


def _import_date_value(value, label):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    raw = str(value).strip()
    if not raw:
        return ""

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass
    raise ValueError(f"{label} doit être une date valide.")


def _import_amount_value(value):
    if value in (None, ""):
        return "0"
    return str(value).strip()


def _admin_project_queryset_from_request(request):
    queryset = TechnicalProject.objects.select_related("categorie").order_by("reference")
    q = (request.GET.get("q") or "").strip()
    if q:
        queryset = queryset.filter(
            Q(reference__icontains=q)
            | Q(affaire__icontains=q)
            | Q(name__icontains=q)
            | Q(adresse_bien__icontains=q)
            | Q(vendeur__icontains=q)
            | Q(beneficiaire__icontains=q)
            | Q(locataire__icontains=q)
        )
    return queryset


def _admin_project_export_columns():
    columns = list(ADMIN_PROJECT_EXPORT_COLUMNS)
    for field in _custom_field_queryset(include_inactive=False).filter(show_in_table=True):
        columns.append((field.label, f"custom:{field.pk}"))
    return columns


def _admin_project_export_row(project, columns):
    serialized = _serialize_project(project)
    row = []
    custom_display = {
        str(item["field_id"]): item["display_value"]
        for item in serialized.get("custom_fields_display", [])
    }
    for _, field in columns:
        if field.startswith("custom:"):
            row.append(custom_display.get(field.split(":", 1)[1], ""))
        else:
            row.append(serialized.get(field, ""))
    return row


def _size_admin_project_sheet(sheet):
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)


def _append_admin_project_sheet(workbook, title, queryset, use_active=False):
    sheet = workbook.active if use_active else workbook.create_sheet(title=title)
    sheet.title = title
    columns = _admin_project_export_columns()
    sheet.append([label for label, _ in columns])
    for project in queryset:
        sheet.append(_admin_project_export_row(project, columns))
    _size_admin_project_sheet(sheet)
    return sheet


def _admin_project_pdf_paragraph(value, style):
    text = escape(str(value or "")).replace("\n", "<br/>")
    return Paragraph(text, style)


def _build_admin_project_pdf(queryset):
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title="Dossiers administratifs",
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    heading_style = styles["Heading2"]
    cell_style = styles["BodyText"]
    cell_style.fontSize = 8
    cell_style.leading = 10

    story = [Paragraph("Dossiers administratifs", title_style), Spacer(1, 0.4 * cm)]
    projects = list(queryset)
    if not projects:
        story.append(Paragraph("Aucun dossier à exporter.", styles["BodyText"]))
    columns = _admin_project_export_columns()
    for project in projects:
        serialized = _serialize_project(project)
        title = f"{serialized.get('reference', '')} - {serialized.get('affaire', '')}".strip(" -")
        story.append(Paragraph(escape(title), heading_style))
        rows = []
        custom_display = {
            str(item["field_id"]): item["display_value"]
            for item in serialized.get("custom_fields_display", [])
        }
        for label, field in columns:
            value = custom_display.get(field.split(":", 1)[1], "") if field.startswith("custom:") else serialized.get(field, "")
            rows.append(
                [
                    Paragraph(f"<b>{escape(label)}</b>", cell_style),
                    _admin_project_pdf_paragraph(value, cell_style),
                ]
            )
        table = Table(rows, colWidths=[5 * cm, 12.5 * cm], hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.extend([table, Spacer(1, 0.45 * cm)])

    document.build(story)
    output.seek(0)
    return output


def _unique_import_reference(row_number):
    candidate = f"ADM-IMP-{int(row_number):04d}"
    if not TechnicalProject.objects.filter(reference=candidate).exists():
        return candidate

    suffix = 1
    while True:
        candidate = f"ADM-IMP-{int(row_number):04d}-{suffix}"
        if not TechnicalProject.objects.filter(reference=candidate).exists():
            return candidate
        suffix += 1


def _import_category_id(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    if raw.isdigit():
        return raw

    _admin_project_categories()
    category = CategorieDossierAdministratif.objects.filter(nom__iexact=raw).first()
    if not category:
        raise ValueError(f"Catégorie de dossier invalide : {raw}")
    return str(category.pk)


def _default_activity_from_sheet(title):
    normalized = _normalize_header(title)
    if "promotion" in normalized:
        return "promotion_immobiliere"
    return "marchand_biens"


def _admin_project_payload_from_import(row, row_number, default_activite_metier="marchand_biens"):
    payload = {}
    custom_fields = {}
    date_fields = {
        "date_promesse",
        "date_dg",
        "date_cs_pret",
        "date_reiteration",
        "depot_permis",
        "obtention_permis",
    }
    amount_fields = {"frais", "prix", "dg"}

    for field, value in row.items():
        if str(field).startswith("custom:"):
            custom_fields[str(field).split(":", 1)[1]] = value
        elif field in date_fields:
            payload[field] = _import_date_value(value, field)
        elif field in amount_fields:
            payload[field] = _import_amount_value(value)
        elif field == "type_dossier":
            payload[field] = _choice_value(value, TechnicalProject.ADMIN_DOSSIER_TYPES, "Type de dossier")
        elif field == "activite_metier":
            payload[field] = _choice_value(value, TechnicalProject.ACTIVITES_METIER, "Activité métier")
        elif field == "etat":
            payload[field] = _choice_value(value, TechnicalProject.ETATS, "État")
        elif field == "categorie":
            payload["categorie_id"] = _import_category_id(value)
        else:
            payload[field] = str(value or "").strip()

    if custom_fields:
        payload["custom_fields"] = custom_fields

    if not payload.get("reference"):
        payload["reference"] = _unique_import_reference(row_number)
    if not payload.get("affaire"):
        raise ValueError("Affaire obligatoire.")
    if not payload.get("type_dossier"):
        payload["type_dossier"] = "vente"
    if not payload.get("activite_metier"):
        payload["activite_metier"] = default_activite_metier
    if not payload.get("etat"):
        payload["etat"] = "promesse"
    if not payload.get("categorie_id"):
        payload["categorie_id"] = str(_default_categorie().pk)

    return payload


def _iter_admin_project_import_rows(uploaded_file):
    filename = (uploaded_file.name or "").lower()
    if filename.endswith(".xlsx"):
        workbook = load_workbook(uploaded_file, data_only=True)
        sheets = [
            (sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)])
            for sheet in workbook.worksheets
        ]
    else:
        raise ValueError("Format non supporté. Merci d’importer un fichier .xlsx.")

    parsed_rows = []
    custom_field_by_header = {
        _normalize_header(field.label): f"custom:{field.pk}"
        for field in _custom_field_queryset(include_inactive=False)
    }
    for sheet_title, rows in sheets:
        if not rows:
            continue

        headers = rows[0]
        field_by_index = {}
        for index, header in enumerate(headers):
            normalized = _normalize_header(header)
            field = ADMIN_PROJECT_IMPORT_ALIASES.get(normalized)
            if not field:
                field = custom_field_by_header.get(normalized)
            if field:
                field_by_index[index] = field

        if not field_by_index:
            continue

        default_activite_metier = _default_activity_from_sheet(sheet_title)
        for row_number, values in enumerate(rows[1:], start=2):
            row_data = {}
            has_value = False
            for index, field in field_by_index.items():
                value = values[index] if index < len(values) else ""
                if value not in (None, ""):
                    has_value = True
                row_data[field] = value
            if has_value:
                parsed_rows.append((len(parsed_rows) + 2, f"{sheet_title} ligne {row_number}", row_data, default_activite_metier))

    if not parsed_rows:
        raise ValueError("Aucune colonne reconnue dans le fichier importé.")
    return parsed_rows


def _import_admin_projects(uploaded_file, user):
    created = 0
    updated = 0
    errors = []

    for sequence, row_label, row_data, default_activite_metier in _iter_admin_project_import_rows(uploaded_file):
        try:
            payload = _admin_project_payload_from_import(row_data, sequence, default_activite_metier)
            reference = payload["reference"].strip().upper()
            existing = TechnicalProject.objects.filter(reference=reference).first()
            project_data = _project_form_data(payload, existing_project=existing)

            with transaction.atomic():
                if existing:
                    for field, value in project_data.items():
                        setattr(existing, field, value)
                    existing.updated_by = user
                    existing.save()
                    _save_custom_field_values(existing, payload.get("custom_fields") or {})
                    updated += 1
                else:
                    project = TechnicalProject.objects.create(
                        created_by=user,
                        updated_by=user,
                        **project_data,
                    )
                    _save_custom_field_values(project, payload.get("custom_fields") or {})
                    created += 1
        except Exception as exc:
            errors.append(f"{row_label} : {exc}")

    return {"created": created, "updated": updated, "errors": errors}


def _next_activity_id():
    numeric_ids = []
    for value in Activite.objects.values_list("id", flat=True):
        if str(value).isdigit():
            numeric_ids.append(int(value))
    return str(max(numeric_ids, default=0) + 1)


def _user_label(user):
    if not user:
        return ""
    full_name = user.get_full_name()
    return full_name or user.get_username()


def _serialize_activity(activity, include_datetime=False):
    date_value = activity.date
    duree_minutes = activity.duree_minutes or 60
    is_overdue = bool(
        date_value
        and date_value < timezone.now()
        and activity.statut not in {"done", "cancelled"}
    )
    dossier_reference = getattr(activity.dossier, "reference", "") or ""
    dossier_nom = (
        getattr(activity.dossier, "affaire", "")
        or getattr(activity.dossier, "name", "")
        or dossier_reference
    )
    payload = {
        "id": activity.id,
        "titre": activity.titre or "",
        "dossier": dossier_reference,
        "dossier_nom": dossier_nom,
        "type": getattr(activity.type, "type", "") or "",
        "date": date_value.strftime("%Y-%m-%d") if date_value else "",
        "commentaire": activity.commentaire or "",
        "statut": activity.statut,
        "statut_label": activity.get_statut_display(),
        "priorite": activity.priorite,
        "priorite_label": activity.get_priorite_display(),
        "duree_minutes": duree_minutes,
        "duree_label": activity.get_duree_minutes_display(),
        "responsable_id": activity.responsable_id or "",
        "responsable_label": _user_label(activity.responsable),
        "is_overdue": is_overdue,
        "status_color": STATUS_COLORS.get(activity.statut, STATUS_COLORS["todo"]),
        "priority_color": PRIORITY_COLORS.get(activity.priorite, PRIORITY_COLORS["normal"]),
        "outlook_synced": bool(activity.outlook_event_id),
    }
    if include_datetime:
        payload.update(
            {
                "time": date_value.strftime("%H:%M") if date_value else "09:00",
                "datetime": date_value.isoformat() if date_value else None,
                "end_time": (date_value + timedelta(minutes=duree_minutes)).strftime("%H:%M") if date_value else "",
                "end_datetime": (date_value + timedelta(minutes=duree_minutes)).isoformat() if date_value else None,
            }
        )
    return payload


def _default_categorie():
    category, _ = CategorieDossierAdministratif.objects.get_or_create(
        nom=CategorieDossierAdministratif.DEFAULT_NOM,
        defaults={"is_default": True},
    )
    CategorieDossierAdministratif.objects.exclude(pk=category.pk).filter(is_default=True).update(is_default=False)
    return category


def _admin_project_categories():
    categories = {}
    for index, nom in enumerate(CategorieDossierAdministratif.CATEGORIES_OFFICIELLES):
        category, _ = CategorieDossierAdministratif.objects.get_or_create(
            nom=nom,
            defaults={"is_default": index == 0},
        )
        categories[nom] = category

    default_category = categories[CategorieDossierAdministratif.DEFAULT_NOM]
    CategorieDossierAdministratif.objects.exclude(pk=default_category.pk).filter(is_default=True).update(
        is_default=False
    )
    if not default_category.is_default:
        default_category.is_default = True
        default_category.save(update_fields=["is_default"])

    return [categories[nom] for nom in CategorieDossierAdministratif.CATEGORIES_OFFICIELLES]


def _custom_field_queryset(include_inactive=False, activite_metier=None):
    queryset = ChampPersonnaliseDossier.objects.all()
    if not include_inactive:
        queryset = queryset.filter(is_active=True)
    if activite_metier is not None:
        queryset = queryset.filter(Q(activite_metier="") | Q(activite_metier=activite_metier))
    return queryset.order_by("sort_order", "label")


def _custom_field_activity_label(field):
    if not field.activite_metier:
        return "Toutes les activités"
    return dict(TechnicalProject.ACTIVITES_METIER).get(field.activite_metier, field.activite_metier)


def _serialize_custom_field(field):
    return {
        "id": field.pk,
        "label": field.label,
        "activite_metier": field.activite_metier,
        "activite_metier_label": _custom_field_activity_label(field),
        "field_type": field.field_type,
        "field_type_label": field.get_field_type_display(),
        "choices": field.choice_list,
        "choices_text": field.choices,
        "show_in_detail": field.show_in_detail,
        "show_in_table": field.show_in_table,
        "is_active": field.is_active,
        "sort_order": field.sort_order,
    }


def _custom_field_value_map(project):
    return {
        str(value.field_id): value.value
        for value in project.custom_field_values.select_related("field").all()
    }


def _custom_field_display_value(field, raw_value):
    if raw_value in (None, ""):
        return ""
    if field.field_type == "checkbox":
        return "Oui" if _truthy(raw_value) else "Non"
    if field.field_type == "amount":
        amount = _parse_non_negative_decimal(raw_value, field.label)
        return f"{amount} €"
    return str(raw_value)


def _custom_field_display_values(project, fields=None):
    fields = fields if fields is not None else _custom_field_queryset(
        include_inactive=True,
        activite_metier=project.activite_metier,
    )
    values = _custom_field_value_map(project)
    displayed = []
    for field in fields:
        raw_value = values.get(str(field.pk), "")
        displayed.append(
            {
                "field_id": field.pk,
                "label": field.label,
                "value": raw_value,
                "display_value": _custom_field_display_value(field, raw_value),
                "show_in_detail": field.show_in_detail,
                "show_in_table": field.show_in_table,
                "is_active": field.is_active,
            }
        )
    return displayed


def _parse_custom_field_value(field, value):
    if field.field_type == "checkbox":
        return "true" if _truthy(value) else "false"

    if field.field_type == "date":
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()

    raw = str(value or "").strip()
    if not raw:
        return ""
    if field.field_type == "date":
        return _parse_iso_date(raw, field.label).isoformat()
    if field.field_type == "amount":
        return str(_parse_non_negative_decimal(raw, field.label))
    if field.field_type == "number":
        try:
            return str(Decimal(raw.replace(",", ".")))
        except (InvalidOperation, ValueError):
            raise ValueError(f"{field.label} doit être un nombre valide.")
    if field.field_type == "choice":
        choices = field.choice_list
        if raw not in choices:
            raise ValueError(f"{field.label} doit correspondre à un choix autorisé.")
    return raw


def _save_custom_field_values(project, values):
    if not isinstance(values, dict):
        raise ValueError("Les champs personnalisés doivent être fournis sous forme d'objet.")

    fields = {
        str(field.pk): field
        for field in _custom_field_queryset(include_inactive=False, activite_metier=project.activite_metier)
    }
    for field_id, raw_value in values.items():
        field = fields.get(str(field_id))
        if not field:
            continue
        value = _parse_custom_field_value(field, raw_value)
        ValeurChampPersonnaliseDossier.objects.update_or_create(
            dossier=project,
            field=field,
            defaults={"value": value},
        )


def _custom_field_form_data(data, existing_field=None):
    label = (data.get("label") or "").strip()
    activite_metier = (data.get("activite_metier") or "").strip()
    field_type = (data.get("field_type") or "text").strip()
    choices = (data.get("choices") or "").strip()

    if not label:
        raise ValueError("Le libellé du champ personnalisé est obligatoire.")
    if activite_metier and activite_metier not in dict(TechnicalProject.ACTIVITES_METIER):
        raise ValueError("Activité métier du champ personnalisé invalide.")
    if field_type not in dict(ChampPersonnaliseDossier.FIELD_TYPES):
        raise ValueError("Type de champ personnalisé invalide.")
    if field_type == "choice" and not [item for item in choices.splitlines() if item.strip()]:
        raise ValueError("Une liste de choix doit contenir au moins une option.")
    if field_type != "choice":
        choices = ""

    duplicate = ChampPersonnaliseDossier.objects.filter(label__iexact=label)
    if existing_field:
        duplicate = duplicate.exclude(pk=existing_field.pk)
    if duplicate.exists():
        raise ValueError("Un champ personnalisé avec ce libellé existe déjà.")

    try:
        sort_order = int(data.get("sort_order") or 0)
    except (TypeError, ValueError):
        raise ValueError("L'ordre d'affichage doit être un entier.")

    return {
        "label": label,
        "activite_metier": activite_metier,
        "field_type": field_type,
        "choices": choices,
        "show_in_detail": _truthy(data.get("show_in_detail", True)),
        "show_in_table": _truthy(data.get("show_in_table", False)),
        "is_active": _truthy(data.get("is_active", True)),
        "sort_order": max(sort_order, 0),
    }


def _serialize_project(project):
    categorie = project.categorie
    custom_values = _custom_field_value_map(project)
    active_custom_fields = list(_custom_field_queryset(include_inactive=False, activite_metier=project.activite_metier))
    return {
        "id": project.pk,
        "reference": project.reference,
        "name": project.affaire or project.name,
        "affaire": project.affaire or project.name,
        "lot_etage": project.lot_etage,
        "adresse_bien": project.adresse_bien,
        "parcelles": project.parcelles,
        "vendeur": project.vendeur,
        "beneficiaire": project.beneficiaire,
        "locataire": project.locataire,
        "type": project.type_dossier,
        "type_label": project.get_type_dossier_display(),
        "type_dossier": project.type_dossier,
        "type_dossier_label": project.get_type_dossier_display(),
        "activite_metier": project.activite_metier,
        "activite_metier_label": project.get_activite_metier_display(),
        "etat": project.etat,
        "etat_label": project.get_etat_display(),
        "categorie_id": categorie.pk if categorie else "",
        "categorie_label": categorie.nom if categorie else "",
        "date_promesse": project.date_promesse.isoformat() if project.date_promesse else "",
        "premiere_periode": project.premiere_periode,
        "deuxieme_periode": project.deuxieme_periode,
        "avenant_1": project.avenant_1,
        "avenant_2": project.avenant_2,
        "avenant_3": project.avenant_3,
        "negociation_externe": project.negociation_externe,
        "frais": str(project.frais),
        "prix": str(project.prix),
        "dg": str(project.dg),
        "date_dg": project.date_dg.isoformat() if project.date_dg else "",
        "depot_permis": project.depot_permis.isoformat() if project.depot_permis else "",
        "obtention_permis": project.obtention_permis.isoformat() if project.obtention_permis else "",
        "diags": project.diags,
        "bornage": project.bornage,
        "etude_sol_geotechnique": project.etude_sol_geotechnique,
        "etude_pollution": project.etude_pollution,
        "etude_impact": project.etude_impact,
        "prorogation": project.prorogation,
        "cs_pret": project.cs_pret,
        "date_cs_pret": project.date_cs_pret.isoformat() if project.date_cs_pret else "",
        "date_reiteration": project.date_reiteration.isoformat() if project.date_reiteration else "",
        "acte": project.acte,
        "releves_compte": project.releves_compte,
        "total_estimated": str(project.total_estimated),
        "activities_count": Activite.objects.filter(dossier=project).count(),
        "custom_fields": custom_values,
        "custom_fields_display": _custom_field_display_values(project, active_custom_fields),
    }


def _project_form_data(data, existing_project=None):
    reference = (data.get("reference") or "").strip().upper()
    affaire = (data.get("affaire") or data.get("name") or "").strip()
    type_dossier = (data.get("type_dossier") or data.get("type") or "vente").strip()
    activite_metier = (data.get("activite_metier") or "marchand_biens").strip()
    etat = (data.get("etat") or "promesse").strip()
    categorie_id = str(data.get("categorie_id") or data.get("categorie") or "").strip()

    if not reference or not affaire:
        raise ValueError("La référence et l'affaire du dossier sont obligatoires.")
    if type_dossier not in dict(TechnicalProject.ADMIN_DOSSIER_TYPES):
        raise ValueError("Type de dossier invalide.")
    if activite_metier not in dict(TechnicalProject.ACTIVITES_METIER):
        raise ValueError("Activité métier invalide.")
    if etat not in dict(TechnicalProject.ETATS):
        raise ValueError("État de dossier invalide.")

    categorie = None
    if categorie_id:
        categorie = CategorieDossierAdministratif.objects.filter(
            pk=categorie_id,
            nom__in=CategorieDossierAdministratif.CATEGORIES_OFFICIELLES,
        ).first()
        if not categorie:
            raise ValueError("Catégorie de dossier invalide.")
    else:
        categorie = _default_categorie()

    frais = _parse_non_negative_decimal(data.get("frais"), "Frais")
    prix = _parse_non_negative_decimal(data.get("prix") or data.get("total_estimated"), "Prix")
    dg = _parse_non_negative_decimal(data.get("dg"), "DG")

    duplicate = TechnicalProject.objects.filter(reference=reference)
    if existing_project:
        duplicate = duplicate.exclude(pk=existing_project.pk)
    if duplicate.exists():
        raise ValueError(f'La référence "{reference}" existe déjà.')

    return {
        "reference": reference,
        "name": affaire,
        "affaire": affaire,
        "lot_etage": (data.get("lot_etage") or "").strip(),
        "adresse_bien": (data.get("adresse_bien") or "").strip(),
        "parcelles": (data.get("parcelles") or "").strip(),
        "vendeur": (data.get("vendeur") or "").strip(),
        "beneficiaire": (data.get("beneficiaire") or "").strip(),
        "locataire": (data.get("locataire") or "").strip(),
        "type_dossier": type_dossier,
        "activite_metier": activite_metier,
        "etat": etat,
        "categorie": categorie,
        "date_promesse": _parse_iso_date(data.get("date_promesse"), "Date de promesse"),
        "premiere_periode": (data.get("premiere_periode") or "").strip(),
        "deuxieme_periode": (data.get("deuxieme_periode") or "").strip(),
        "avenant_1": (data.get("avenant_1") or "").strip(),
        "avenant_2": (data.get("avenant_2") or "").strip(),
        "avenant_3": (data.get("avenant_3") or "").strip(),
        "negociation_externe": (data.get("negociation_externe") or "").strip(),
        "frais": frais,
        "prix": prix,
        "dg": dg,
        "date_dg": _parse_iso_date(data.get("date_dg"), "Date DG"),
        "depot_permis": _parse_iso_date(data.get("depot_permis"), "Dépôt permis"),
        "obtention_permis": _parse_iso_date(data.get("obtention_permis"), "Obtention permis"),
        "diags": (data.get("diags") or "").strip(),
        "bornage": (data.get("bornage") or "").strip(),
        "etude_sol_geotechnique": (data.get("etude_sol_geotechnique") or "").strip(),
        "etude_pollution": (data.get("etude_pollution") or "").strip(),
        "etude_impact": (data.get("etude_impact") or "").strip(),
        "prorogation": (data.get("prorogation") or "").strip(),
        "cs_pret": (data.get("cs_pret") or "").strip(),
        "date_cs_pret": _parse_iso_date(data.get("date_cs_pret"), "Date CS prêt"),
        "date_reiteration": _parse_iso_date(data.get("date_reiteration"), "Date de réitération"),
        "acte": (data.get("acte") or "").strip(),
        "releves_compte": (data.get("releves_compte") or "").strip(),
        "total_estimated": prix,
    }


def _project_delete_blockers(project):
    blockers = []
    activities_count = Activite.objects.filter(dossier=project).count()
    related_counts = {
        "document(s)": project.documents.count(),
        "date(s) clé(s)": project.key_dates.count(),
        "action(s) technique(s)": project.actions.count(),
        "facture(s)": Facture.objects.filter(dossier=project).count(),
        "dépense(s)": project.expenses.count(),
        "e-mail(s) technique(s)": project.emails.count(),
        "entrée(s) d'historique": project.history.count(),
    }

    if activities_count:
        blockers.append(f"{activities_count} activité(s)")
    for label, count in related_counts.items():
        if count:
            blockers.append(f"{count} {label}")
    return blockers


def _active_activity_reminder_rules():
    return list(RegleRappelActivite.objects.filter(is_active=True).order_by("timing", "days"))


def _apply_calendar_filters(queryset, params):
    type_label = (params.get("type") or "").strip()
    dossier_ref = (params.get("dossier") or "").strip()
    responsable = (params.get("responsable") or "").strip()
    statut = (params.get("statut") or "").strip()
    priorite = (params.get("priorite") or "").strip()
    client = (params.get("client") or "").strip()
    contact = (params.get("contact") or "").strip()
    date_from = (params.get("date_from") or "").strip()
    date_to = (params.get("date_to") or "").strip()

    if type_label:
        queryset = queryset.filter(type__type__iexact=type_label)
    if dossier_ref:
        queryset = queryset.filter(dossier__reference=dossier_ref)
    if responsable:
        queryset = queryset.filter(responsable_id=responsable)
    if statut:
        queryset = queryset.filter(statut=statut)
    if priorite:
        queryset = queryset.filter(priorite=priorite)
    if date_from:
        queryset = queryset.filter(date__date__gte=datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        queryset = queryset.filter(date__date__lte=datetime.strptime(date_to, "%Y-%m-%d").date())
    return queryset


def _ics_escape(value):
    return (
        str(value or "")
        .replace("\\", "\\\\")
        .replace(";", "\\;")
        .replace(",", "\\,")
        .replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("\n", "\\n")
    )


def _ics_datetime(value):
    if timezone.is_naive(value):
        value = timezone.make_aware(value, timezone.get_current_timezone())
    return value.astimezone(dt_timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _fold_ics_line(line):
    if len(line) <= 75:
        return [line]
    parts = [line[:75]]
    rest = line[75:]
    while rest:
        parts.append(" " + rest[:74])
        rest = rest[74:]
    return parts


def _ics_line(name, value):
    return _fold_ics_line(f"{name}:{value}")


def _unfold_ics_lines(content):
    lines = content.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    unfolded = []
    for line in lines:
        if line.startswith((" ", "\t")) and unfolded:
            unfolded[-1] += line[1:]
        else:
            unfolded.append(line)
    return unfolded


def _ics_unescape(value):
    return (
        str(value or "")
        .replace("\\n", "\n")
        .replace("\\N", "\n")
        .replace("\\,", ",")
        .replace("\\;", ";")
        .replace("\\\\", "\\")
    )


def _parse_ics_property(line):
    if ":" not in line:
        return None, {}, ""
    name_and_params, value = line.split(":", 1)
    parts = name_and_params.split(";")
    name = parts[0].upper()
    params = {}
    for part in parts[1:]:
        if "=" in part:
            key, param_value = part.split("=", 1)
            params[key.upper()] = param_value.strip('"')
    return name, params, value


def _parse_ics_datetime(value, params):
    raw = str(value or "").strip()
    if not raw:
        raise ValueError("Date d'événement manquante.")

    if params.get("VALUE", "").upper() == "DATE" or re.fullmatch(r"\d{8}", raw):
        parsed_date = datetime.strptime(raw[:8], "%Y%m%d").date()
        return timezone.make_aware(datetime.combine(parsed_date, datetime.min.time()).replace(hour=9))

    is_utc = raw.endswith("Z")
    if is_utc:
        raw = raw[:-1]

    try:
        parsed = datetime.strptime(raw, "%Y%m%dT%H%M%S")
    except ValueError:
        parsed = datetime.strptime(raw, "%Y%m%dT%H%M")

    if is_utc:
        return parsed.replace(tzinfo=dt_timezone.utc).astimezone(timezone.get_current_timezone())
    return timezone.make_aware(parsed, timezone.get_current_timezone())


def _parse_ics_duration_minutes(value):
    raw = str(value or "").strip().upper()
    if not raw:
        return None
    match = re.fullmatch(r"P(?:(?P<days>\d+)D)?(?:T(?:(?P<hours>\d+)H)?(?:(?P<minutes>\d+)M)?(?:(?P<seconds>\d+)S)?)?", raw)
    if not match:
        return None
    days = int(match.group("days") or 0)
    hours = int(match.group("hours") or 0)
    minutes = int(match.group("minutes") or 0)
    seconds = int(match.group("seconds") or 0)
    total = days * 1440 + hours * 60 + minutes + (1 if seconds else 0)
    return total or None


def _normalize_import_duration(minutes):
    allowed = [value for value, _ in Activite.DUREES_CRENEAU]
    if not minutes:
        return 60
    return min(allowed, key=lambda value: abs(value - minutes))


def _parse_calendar_ics_events(uploaded_file):
    filename = (uploaded_file.name or "").lower()
    if not filename.endswith(".ics"):
        raise ValueError("Format non supporté. Merci d'importer un fichier .ics.")

    raw = uploaded_file.read()
    try:
        content = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = raw.decode("latin-1")

    events = []
    current = None
    for line in _unfold_ics_lines(content):
        stripped = line.strip()
        if stripped == "BEGIN:VEVENT":
            current = {}
            continue
        if stripped == "END:VEVENT":
            if current is not None:
                events.append(current)
            current = None
            continue
        if current is None:
            continue

        name, params, value = _parse_ics_property(line)
        if not name:
            continue
        if name in {"SUMMARY", "DESCRIPTION", "UID"}:
            current[name] = _ics_unescape(value).strip()
        elif name in {"DTSTART", "DTEND"}:
            current[name] = _parse_ics_datetime(value, params)
        elif name == "DURATION":
            current[name] = _parse_ics_duration_minutes(value)

    parsed_events = []
    for event in events:
        start = event.get("DTSTART")
        if not start:
            continue
        end = event.get("DTEND")
        duration = event.get("DURATION")
        if end:
            duration = max(int((end - start).total_seconds() // 60), 1)
        duration = _normalize_import_duration(duration)
        parsed_events.append(
            {
                "uid": event.get("UID", ""),
                "titre": event.get("SUMMARY") or "Événement importé",
                "commentaire": event.get("DESCRIPTION", ""),
                "date": start,
                "duree_minutes": duration,
            }
        )

    if not parsed_events:
        raise ValueError("Aucun événement calendrier exploitable n'a été trouvé.")
    return parsed_events


def _import_calendar_ics(uploaded_file, dossier, type_activite, responsable, user):
    events = _parse_calendar_ics_events(uploaded_file)
    created = 0
    skipped = 0

    for event in events:
        duplicate = Activite.objects.filter(
            dossier=dossier,
            type=type_activite,
            titre=event["titre"],
            date=event["date"],
        ).exists()
        if duplicate:
            skipped += 1
            continue

        commentaire_parts = []
        if event["commentaire"]:
            commentaire_parts.append(event["commentaire"])
        if event["uid"]:
            commentaire_parts.append(f"UID calendrier : {event['uid']}")

        Activite.objects.create(
            id=_next_activity_id(),
            titre=event["titre"],
            dossier=dossier,
            type=type_activite,
            date=event["date"],
            duree_minutes=event["duree_minutes"],
            date_type="date",
            commentaire="\n\n".join(commentaire_parts),
            statut="todo",
            priorite="normal",
            responsable=responsable,
            created_by=user,
            updated_by=user,
        )
        created += 1

    return {"created": created, "skipped": skipped, "total": len(events)}


def _calendar_export_queryset_from_request(request):
    queryset = (
        Activite.objects.filter(date__isnull=False)
        .select_related("dossier", "type", "responsable")
        .order_by("date", "id")
    )

    month_raw = (request.GET.get("month") or "").strip()
    year_raw = (request.GET.get("year") or "").strip()
    if month_raw or year_raw:
        try:
            month = int(month_raw)
            year = int(year_raw)
            if month < 1 or month > 12:
                raise ValueError
        except ValueError:
            raise ValueError("Mois ou année invalide.")

        start_date = timezone.make_aware(datetime(year, month, 1))
        end_date = (
            timezone.make_aware(datetime(year + 1, 1, 1))
            if month == 12
            else timezone.make_aware(datetime(year, month + 1, 1))
        )
        queryset = queryset.filter(date__gte=start_date, date__lt=end_date)

    return _apply_calendar_filters(queryset, request.GET)


def _build_calendar_ics(activities):
    now_stamp = _ics_datetime(timezone.now())
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Benjamin Immobilier//Intranet Administratif//FR",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:Calendrier administratif",
    ]

    for activity in activities:
        start = activity.date
        duration = activity.duree_minutes or 60
        end = start + timedelta(minutes=duration)
        dossier_label = getattr(activity.dossier, "affaire", "") or getattr(activity.dossier, "name", "") or ""
        summary = activity.titre or f"{activity.type} - {activity.dossier.reference}"
        description_parts = [
            f"Dossier : {activity.dossier.reference} {dossier_label}".strip(),
            f"Type : {activity.type}",
            f"Statut : {activity.get_statut_display()}",
            f"Priorité : {activity.get_priorite_display()}",
            f"Créneau : {activity.get_duree_minutes_display()}",
        ]
        if activity.responsable:
            description_parts.append(f"Responsable : {_user_label(activity.responsable)}")
        if activity.commentaire:
            description_parts.extend(["", activity.commentaire])

        lines.append("BEGIN:VEVENT")
        for folded in _ics_line("UID", _ics_escape(f"activite-{activity.pk}@benjamin-intranet")):
            lines.append(folded)
        lines.append(f"DTSTAMP:{now_stamp}")
        lines.append(f"DTSTART:{_ics_datetime(start)}")
        lines.append(f"DTEND:{_ics_datetime(end)}")
        for folded in _ics_line("SUMMARY", _ics_escape(summary)):
            lines.append(folded)
        for folded in _ics_line("DESCRIPTION", _ics_escape("\n".join(description_parts))):
            lines.append(folded)
        lines.append("END:VEVENT")

    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"


def _activity_form_data(data):
    dossier_ref = (data.get("dossier") or "").strip()
    type_label = (data.get("type") or "").strip()
    date_str = (data.get("date") or "").strip()

    if not dossier_ref or not type_label or not date_str:
        raise ValueError("Champs obligatoires manquants")

    dossier_obj = TechnicalProject.objects.filter(reference=dossier_ref).first()
    if not dossier_obj:
        raise LookupError(f'Dossier introuvable : "{dossier_ref}"')

    type_obj = TypeActivite.objects.filter(type__iexact=type_label).first()
    if not type_obj:
        types_disponibles = ", ".join(
            TypeActivite.objects.order_by("type").values_list("type", flat=True)
        )
        raise ValueError(f'Type invalide. Types autorisés : {types_disponibles}')

    responsable_id = (data.get("responsable") or "").strip()
    responsable = None
    if responsable_id:
        responsable = Utilisateur.objects.filter(pk=responsable_id, is_active=True).first()
        if not responsable:
            raise ValueError("Responsable introuvable")

    statut = (data.get("statut") or "todo").strip()
    priorite = (data.get("priorite") or "normal").strip()
    duree_raw = str(data.get("duree_minutes") or "60").strip()
    if statut not in dict(Activite.STATUTS):
        raise ValueError("Statut invalide")
    if priorite not in dict(Activite.PRIORITES):
        raise ValueError("Priorité invalide")
    try:
        duree_minutes = int(duree_raw)
    except ValueError:
        raise ValueError("Durée de créneau invalide.")
    if duree_minutes not in dict(Activite.DUREES_CRENEAU):
        raise ValueError("Durée de créneau invalide.")

    return {
        "titre": (data.get("titre") or "").strip(),
        "dossier": dossier_obj,
        "type": type_obj,
        "date": _parse_iso_datetime(date_str),
        "duree_minutes": duree_minutes,
        "date_type": "date",
        "commentaire": (data.get("commentaire") or "").strip() or None,
        "statut": statut,
        "priorite": priorite,
        "responsable": responsable,
    }


def _sync_outlook_after_save(request, activity, sync_requested):
    if sync_requested:
        result = (
            update_outlook_event(request.user, activity)
            if activity.outlook_event_id
            else create_outlook_event(request.user, activity)
        )
        if result.get("success"):
            event_id = result.get("event_id")
            if event_id and activity.outlook_event_id != event_id:
                activity.outlook_event_id = event_id
                activity.save(update_fields=["outlook_event_id"])
            return ""
        return result.get("message") or "Synchronisation Outlook impossible."

    if activity.outlook_event_id:
        event_id = activity.outlook_event_id
        result = delete_outlook_event(request.user, event_id)
        activity.outlook_event_id = ""
        activity.save(update_fields=["outlook_event_id"])
        if not result.get("success"):
            return result.get("message") or "L'événement Outlook n'a pas pu être supprimé."
    return ""


@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def administratif_view(request):
    """
    Page du pôle administratif.
    Affiche le journal persistant des conversations Gmail.
    """
    user = request.user

    conversations = GmailConversation.objects.filter(owner=user).prefetch_related("events")
    journal_status = (request.GET.get("journal_status") or "").strip()
    journal_q = (request.GET.get("journal_q") or "").strip()
    if journal_status:
        conversations = conversations.filter(status=journal_status)
    if journal_q:
        conversations = conversations.filter(
            Q(subject__icontains=journal_q)
            | Q(recipient__icontains=journal_q)
            | Q(preview__icontains=journal_q)
        )
    conversations = conversations[:100]

    dossiers = TechnicalProject.objects.all().order_by("reference")
    types = TypeActivite.objects.all().order_by("type")
    users = Utilisateur.objects.filter(is_active=True).order_by(
        "last_name",
        "first_name",
        "username",
    )
    notifications = list(
        NotificationInterne.objects.filter(user=user, is_read=False)
        .select_related("activite")
        .order_by("-created_at")[:8]
    )
    reminder_rules = _active_activity_reminder_rules()
    last_journal_sync = (
        GmailConversation.objects.filter(owner=user, last_synced_at__isnull=False)
        .order_by("-last_synced_at")
        .values_list("last_synced_at", flat=True)
        .first()
    )

    return render(
        request,
        "management.html",
        {
            "pole_name": "Administratif",
            "conversations": conversations,
            "journal_status": journal_status,
            "journal_q": journal_q,
            "journal_status_choices": GmailConversation.STATUS_CHOICES,
            "last_journal_sync": last_journal_sync,
            "dossiers": dossiers,
            "types": types,
            "users": users,
            "notifications": notifications,
            "notification_count": len(notifications),
            "reminder_rules": reminder_rules,
            "reminder_timing_choices": RegleRappelActivite.TIMING_CHOICES,
        },
    )


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def create_activity_reminder_rule_view(request):
    timing = (request.POST.get("timing") or "").strip()
    days_raw = (request.POST.get("days") or "").strip()

    if timing not in dict(RegleRappelActivite.TIMING_CHOICES):
        messages.error(request, "Sens du rappel invalide.")
        return redirect("admin_view")

    try:
        days = int(days_raw)
    except ValueError:
        messages.error(request, "Le nombre de jours doit être un entier.")
        return redirect("admin_view")

    if days < 0 or days > 365:
        messages.error(request, "Le nombre de jours doit être compris entre 0 et 365.")
        return redirect("admin_view")
    if days == 0:
        timing = "before"

    rule, created = RegleRappelActivite.objects.get_or_create(
        timing=timing,
        days=days,
        defaults={"is_active": True},
    )
    if not created and not rule.is_active:
        rule.is_active = True
        rule.save(update_fields=["is_active"])
        messages.success(request, f"Règle {rule.label} réactivée.")
    elif created:
        messages.success(request, f"Règle {rule.label} ajoutée.")
    else:
        messages.info(request, f"La règle {rule.label} existe déjà.")
    return redirect("admin_view")


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def delete_activity_reminder_rule_view(request, rule_id):
    rule = get_object_or_404(RegleRappelActivite, pk=rule_id)
    label = rule.label
    rule.delete()
    messages.success(request, f"Règle {label} supprimée.")
    return redirect("admin_view")


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def sync_gmail_journal_view(request):
    """
    Synchronise le journal Gmail à la demande, sans bloquer le rendu de la page.
    """
    try:
        result = sync_conversation_journal(request.user, limit=100)
        return JsonResponse({"success": True, **result})
    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)


@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def admin_dossiers_view(request):
    categories = _admin_project_categories()
    custom_fields = list(_custom_field_queryset(include_inactive=True))
    q = (request.GET.get("q") or "").strip()
    dossiers = _admin_project_queryset_from_request(request)
    return render(
        request,
        "admin_projects.html",
        {
            "pole_name": "Administratif",
            "projets": [_serialize_project(dossier) for dossier in dossiers],
            "dossier_type_choices": TechnicalProject.ADMIN_DOSSIER_TYPES,
            "activite_metier_choices": TechnicalProject.ACTIVITES_METIER,
            "etat_choices": TechnicalProject.ETATS,
            "categories": categories,
            "custom_fields": [_serialize_custom_field(field) for field in custom_fields],
            "search_query": q,
        },
    )


@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def admin_dossiers_export_view(request):
    queryset = _admin_project_queryset_from_request(request)

    workbook = Workbook()
    _append_admin_project_sheet(
        workbook,
        ADMIN_PROJECT_PROMOTION_SHEET,
        queryset.filter(activite_metier="promotion_immobiliere"),
        use_active=True,
    )
    _append_admin_project_sheet(
        workbook,
        ADMIN_PROJECT_OTHER_SHEET,
        queryset.exclude(activite_metier="promotion_immobiliere"),
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dossiers_administratifs.xlsx"'
    return response


@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def admin_dossiers_export_pdf_view(request):
    queryset = _admin_project_queryset_from_request(request)
    output = _build_admin_project_pdf(queryset)
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dossiers_administratifs.pdf"'
    return response


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def admin_dossiers_import_view(request):
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        messages.error(request, "Merci de sélectionner un fichier .xlsx à importer.")
        return redirect("admin_dossiers")

    try:
        result = _import_admin_projects(uploaded_file, request.user)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("admin_dossiers")

    if result["created"] or result["updated"]:
        messages.success(
            request,
            f"Import terminé : {result['created']} dossier(s) créé(s), "
            f"{result['updated']} dossier(s) mis à jour.",
        )
    else:
        messages.warning(request, "Aucun dossier n’a été importé.")

    if result["errors"]:
        preview = " | ".join(result["errors"][:5])
        extra = f" ({len(result['errors']) - 5} erreur(s) supplémentaire(s))" if len(result["errors"]) > 5 else ""
        messages.error(request, f"Certaines lignes n’ont pas été importées : {preview}{extra}")

    return redirect("admin_dossiers")


@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def admin_projects_view(request):
    return redirect("/administratif/dossiers/")


@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def admin_dossier_detail_view(request, dossier_id):
    dossier = get_object_or_404(
        TechnicalProject.objects.select_related("categorie").prefetch_related("custom_field_values__field"),
        pk=dossier_id,
    )
    activites = (
        Activite.objects.filter(dossier=dossier)
        .select_related("type", "responsable")
        .order_by("date", "id")
    )
    return render(
        request,
        "admin_dossier_detail.html",
        {
            "pole_name": "Administratif",
            "dossier": dossier,
            "activites": activites,
            "custom_fields_display": [
                item
                for item in _custom_field_display_values(
                    dossier,
                    _custom_field_queryset(include_inactive=False, activite_metier=dossier.activite_metier),
                )
                if item["show_in_detail"]
            ],
        },
    )


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def send_reply_view(request):
    """
    Envoie une relance dans une conversation Gmail persistante.
    """
    try:
        data = _parse_request_json(request)

        conversation_id = data.get("conversation_id")
        message_text = (data.get("message") or "").strip()

        if not conversation_id:
            return _json_error("Conversation manquante")

        if not message_text:
            return _json_error("Message manquant")

        conversation = get_object_or_404(
            GmailConversation,
            pk=conversation_id,
            owner=request.user,
        )
        result = send_conversation_reminder(
            conversation=conversation,
            user=request.user,
            body=message_text,
        )

        return JsonResponse(result, status=200 if result.get("success") else 400)

    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def generate_auto_message_view(request):
    try:
        data = json.loads(request.body)
        conversation_id = data.get("conversation_id")

        if not conversation_id:
            return JsonResponse({"success": False, "message": "Conversation manquante"}, status=400)

        conversation = get_object_or_404(
            GmailConversation,
            pk=conversation_id,
            owner=request.user,
        )
        destinataire_email = conversation.recipient

        if not destinataire_email:
            return JsonResponse({'success': False, 'message': 'Destinataire introuvable'})

        # Chercher le modèle de relance via l'email client
        from management.models import EmailClient, ModeleRelance, DefaultModeleRelance
        emails = EmailClient.objects.filter(email=destinataire_email)

        if not emails.exists():
            return JsonResponse({'success': False, 'message': 'Client introuvable'})

        metier = emails.first().metier

        try:
            message_relance = ModeleRelance.objects.get(
                utilisateur=request.user.id, metier=metier
            ).message
        except ModeleRelance.DoesNotExist:
            try:
                message_relance = DefaultModeleRelance.objects.get(metier=metier).message
            except DefaultModeleRelance.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': f'Pas de modèle de relance pour le métier {metier}'
                })

        return JsonResponse({'success': True, 'message': message_relance})

    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'}, status=500)


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def update_gmail_conversation_status(request, conversation_id):
    conversation = get_object_or_404(
        GmailConversation,
        pk=conversation_id,
        owner=request.user,
    )
    data = _parse_request_json(request)
    status = (data.get("status") or "").strip()
    if status not in dict(GmailConversation.STATUS_CHOICES):
        return _json_error("Statut invalide")
    if conversation.status == "replied" and status != "replied":
        return _json_error(
            "Une réponse Gmail a été détectée : le statut répondu est prioritaire.",
            status=409,
        )

    old_status = conversation.status
    conversation.status = status
    conversation.save(update_fields=["status", "updated_at"])
    GmailConversationEvent.objects.create(
        conversation=conversation,
        event_type="status_changed",
        user=request.user,
        old_status=old_status,
        new_status=status,
    )
    return JsonResponse({"success": True, "status": status})


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def add_gmail_conversation_note(request, conversation_id):
    conversation = get_object_or_404(
        GmailConversation,
        pk=conversation_id,
        owner=request.user,
    )
    note = (_parse_request_json(request).get("note") or "").strip()
    if not note:
        return _json_error("La note ne peut pas être vide.")
    event = GmailConversationEvent.objects.create(
        conversation=conversation,
        event_type="note",
        user=request.user,
        note=note,
    )
    return JsonResponse(
        {
            "success": True,
            "note": event.note,
            "created_at": timezone.localtime(event.created_at).strftime("%d/%m/%Y %H:%M"),
            "user": request.user.get_full_name() or request.user.username,
        }
    )


@require_http_methods(["GET"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def get_calendar_activities(request):
    """
    Récupère les activités du calendrier pour une vue mensuelle.
    """
    try:
        now = datetime.now()
        month = int(request.GET.get("month", now.month))
        year = int(request.GET.get("year", now.year))

        start_date = timezone.make_aware(datetime(year, month, 1))
        end_date = (
            timezone.make_aware(datetime(year + 1, 1, 1))
            if month == 12
            else timezone.make_aware(datetime(year, month + 1, 1))
        )

        activites = (
            Activite.objects.filter(
                date__gte=start_date,
                date__lt=end_date,
            )
            .select_related("dossier", "type", "responsable")
            .order_by("date", "id")
        )
        activites = _apply_calendar_filters(activites, request.GET)

        activites_list = [_serialize_activity(act, include_datetime=True) for act in activites]

        return JsonResponse(
            {
                "success": True,
                "activites": activites_list,
                "month": month,
                "year": year,
            }
        )

    except ValueError as e:
        return _json_error(f"Paramètre invalide : {str(e)}")
    except Exception as e:
        traceback.print_exc()
        return _json_error(str(e), status=500)


@require_http_methods(["GET"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def get_calendar_activities_week(request):
    """
    Récupère les activités d'une semaine (vue agenda).
    """
    try:
        date_str = request.GET.get("date")
        ref_date = (
            datetime.strptime(date_str, "%Y-%m-%d").date()
            if date_str
            else datetime.now().date()
        )

        monday = ref_date - timedelta(days=ref_date.weekday())
        sunday = monday + timedelta(days=6)

        start_dt = timezone.make_aware(datetime(monday.year, monday.month, monday.day, 0, 0, 0))
        end_dt = timezone.make_aware(datetime(sunday.year, sunday.month, sunday.day, 23, 59, 59))

        activites = (
            Activite.objects.filter(
                date__gte=start_dt,
                date__lte=end_dt,
            )
            .select_related("dossier", "type", "responsable")
            .order_by("date", "id")
        )
        activites = _apply_calendar_filters(activites, request.GET)

        activites_list = [_serialize_activity(act, include_datetime=True) for act in activites]

        return JsonResponse(
            {
                "success": True,
                "activites": activites_list,
                "week_start": monday.isoformat(),
                "week_end": sunday.isoformat(),
            }
        )

    except ValueError as e:
        return _json_error(f"Date invalide : {str(e)}")
    except Exception as e:
        traceback.print_exc()
        return _json_error(str(e), status=500)


@require_http_methods(["GET"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def export_calendar_ics_view(request):
    try:
        activities = list(_calendar_export_queryset_from_request(request))
    except ValueError as e:
        return _json_error(str(e))

    content = _build_calendar_ics(activities)
    response = HttpResponse(content, content_type="text/calendar; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="calendrier_administratif.ics"'
    return response


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def import_calendar_ics_view(request):
    uploaded_file = request.FILES.get("file")
    dossier_ref = (request.POST.get("dossier") or "").strip()
    type_label = (request.POST.get("type") or "").strip()
    responsable_id = (request.POST.get("responsable") or "").strip()

    if not uploaded_file:
        messages.error(request, "Merci de sélectionner un fichier .ics à importer.")
        return redirect("admin_view")
    if not dossier_ref or not type_label:
        messages.error(request, "Le dossier et le type d'activité sont obligatoires pour l'import calendrier.")
        return redirect("admin_view")

    dossier = TechnicalProject.objects.filter(reference=dossier_ref).first()
    if not dossier:
        messages.error(request, "Dossier administratif introuvable pour l'import calendrier.")
        return redirect("admin_view")

    type_activite = TypeActivite.objects.filter(type__iexact=type_label).first()
    if not type_activite:
        messages.error(request, "Type d'activité introuvable pour l'import calendrier.")
        return redirect("admin_view")

    responsable = None
    if responsable_id:
        responsable = Utilisateur.objects.filter(pk=responsable_id, is_active=True).first()
        if not responsable:
            messages.error(request, "Responsable introuvable pour l'import calendrier.")
            return redirect("admin_view")

    try:
        result = _import_calendar_ics(uploaded_file, dossier, type_activite, responsable, request.user)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("admin_view")

    if result["created"]:
        messages.success(
            request,
            f"Import calendrier terminé : {result['created']} activité(s) créée(s), "
            f"{result['skipped']} doublon(s) ignoré(s).",
        )
    else:
        messages.warning(request, f"Aucune activité créée : {result['skipped']} doublon(s) ignoré(s).")
    return redirect("admin_view")


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def create_activity_view(request):
    """
    Crée une nouvelle activité dans le calendrier.
    """
    try:
        data = _parse_request_json(request)
        activity_data = _activity_form_data(data)
        if not activity_data["titre"]:
            activity_data["titre"] = f"{activity_data['type'].type} - {activity_data['dossier'].reference}"

        nouvelle_activite = Activite.objects.create(
            id=_next_activity_id(),
            created_by=request.user,
            updated_by=request.user,
            **activity_data,
        )
        warning = ""
        if "sync_outlook" in data:
            warning = _sync_outlook_after_save(
                request,
                nouvelle_activite,
                _truthy(data.get("sync_outlook")),
            )

        return JsonResponse(
            {
                "success": True,
                "message": "Activité créée avec succès",
                "warning": warning,
                "activity_id": nouvelle_activite.id,
                "activity": _serialize_activity(nouvelle_activite, include_datetime=True),
            }
        )

    except LookupError as e:
        return _json_error(str(e), status=404)
    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def update_activity_view(request, activity_id):
    """
    Met à jour une activité existante depuis le calendrier.
    """
    try:
        data = _parse_request_json(request)
        activity = (
            Activite.objects.select_related("dossier", "type", "responsable")
            .filter(pk=activity_id)
            .first()
        )
        if not activity:
            return _json_error("Activité introuvable", status=404)

        activity_data = _activity_form_data(data)
        if not activity_data["titre"]:
            activity_data["titre"] = f"{activity_data['type'].type} - {activity_data['dossier'].reference}"

        for field, value in activity_data.items():
            setattr(activity, field, value)
        activity.updated_by = request.user
        activity.save()

        warning = ""
        if "sync_outlook" in data:
            warning = _sync_outlook_after_save(
                request,
                activity,
                _truthy(data.get("sync_outlook")),
            )

        activity.refresh_from_db()
        return JsonResponse(
            {
                "success": True,
                "message": "Activité mise à jour avec succès",
                "warning": warning,
                "activity": _serialize_activity(activity, include_datetime=True),
            }
        )

    except LookupError as e:
        return _json_error(str(e), status=404)
    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def delete_activity_view(request):
    """
    Supprime une ou plusieurs activités correspondant aux critères donnés.
    """
    try:
        data = _parse_request_json(request)
        activity_id = (data.get("activity_id") or data.get("id") or "").strip()

        if activity_id:
            activity = Activite.objects.filter(pk=activity_id).first()
            if not activity:
                return _json_error("Activité introuvable", status=404)

            warning = ""
            if activity.outlook_event_id:
                result = delete_outlook_event(request.user, activity.outlook_event_id)
                if not result.get("success"):
                    warning = result.get("message") or "L'événement Outlook n'a pas pu être supprimé."
            activity.delete()
            return JsonResponse(
                {
                    "success": True,
                    "message": "Activité supprimée avec succès",
                    "warning": warning,
                    "deleted_count": 1,
                }
            )

        dossier_ref = (data.get("dossier") or "").strip()
        type_label = (data.get("type") or "").strip()
        date_str = (data.get("date") or "").strip()

        if not dossier_ref or not type_label or not date_str:
            return _json_error("Champs obligatoires manquants")

        dossier_obj = TechnicalProject.objects.filter(reference=dossier_ref).first()
        if not dossier_obj:
            return _json_error(f'Dossier introuvable : "{dossier_ref}"', status=404)

        type_obj = TypeActivite.objects.filter(type__iexact=type_label).first()
        if not type_obj:
            return _json_error(f'Type introuvable : "{type_label}"', status=404)

        date_activite = _parse_iso_datetime(date_str)
        date_debut = date_activite.replace(second=0, microsecond=0)
        date_fin = date_debut + timedelta(minutes=1)

        queryset = Activite.objects.filter(
            dossier=dossier_obj,
            type=type_obj,
            date__gte=date_debut,
            date__lt=date_fin,
        )

        count_before = queryset.count()
        if count_before == 0:
            return _json_error(
                "Aucune activité ne correspond à ces critères",
                status=404,
            )

        warning = ""
        for activity in queryset:
            if activity.outlook_event_id:
                result = delete_outlook_event(request.user, activity.outlook_event_id)
                if not result.get("success") and not warning:
                    warning = result.get("message") or "Un événement Outlook n'a pas pu être supprimé."

        deleted_count, _ = queryset.delete()

        return JsonResponse(
            {
                "success": True,
                "message": f"{deleted_count} activité(s) supprimée(s) avec succès",
                "warning": warning,
                "deleted_count": deleted_count,
            }
        )

    except ValueError as e:
        return _json_error(f"Format de date invalide : {str(e)}")
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def mark_notification_read_view(request, notification_id):
    updated = NotificationInterne.objects.filter(
        pk=notification_id,
        user=request.user,
    ).update(is_read=True)
    if not updated:
        return _json_error("Notification introuvable", status=404)
    return JsonResponse({"success": True})


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def create_project_view(request):
    try:
        data = _parse_request_json(request)
        project_data = _project_form_data(data)
        with transaction.atomic():
            project = TechnicalProject.objects.create(
                created_by=request.user,
                updated_by=request.user,
                **project_data,
            )
            _save_custom_field_values(project, data.get("custom_fields") or {})
        return JsonResponse(
            {
                "success": True,
                "message": "Dossier créé avec succès.",
                "project": _serialize_project(project),
                "dossier": _serialize_project(project),
            }
        )
    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def update_project_view(request, project_id):
    try:
        data = _parse_request_json(request)
        project = TechnicalProject.objects.filter(pk=project_id).first()
        if not project:
            return _json_error("Dossier introuvable", status=404)

        project_data = _project_form_data(data, existing_project=project)
        with transaction.atomic():
            for field, value in project_data.items():
                setattr(project, field, value)
            project.updated_by = request.user
            project.save()
            _save_custom_field_values(project, data.get("custom_fields") or {})
        return JsonResponse(
            {
                "success": True,
                "message": "Dossier mis à jour avec succès.",
                "project": _serialize_project(project),
                "dossier": _serialize_project(project),
            }
        )
    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def create_custom_field_view(request):
    try:
        field = ChampPersonnaliseDossier.objects.create(
            **_custom_field_form_data(_parse_request_json(request))
        )
        return JsonResponse({"success": True, "field": _serialize_custom_field(field)})
    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def update_custom_field_view(request, field_id):
    try:
        field = get_object_or_404(ChampPersonnaliseDossier, pk=field_id)
        field_data = _custom_field_form_data(_parse_request_json(request), existing_field=field)
        for name, value in field_data.items():
            setattr(field, name, value)
        field.save()
        return JsonResponse({"success": True, "field": _serialize_custom_field(field)})
    except ValueError as e:
        return _json_error(str(e))
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)


@require_http_methods(["POST"])
@login_required
@user_passes_test(has_administratif_access, login_url="/", redirect_field_name=None)
def delete_project_view(request, project_id):
    try:
        project = TechnicalProject.objects.filter(pk=project_id).first()
        if not project:
            return _json_error("Dossier introuvable", status=404)

        blockers = _project_delete_blockers(project)
        if blockers:
            return _json_error(
                "Suppression impossible : ce dossier est lié à "
                + ", ".join(blockers)
                + ".",
                status=409,
            )

        project.delete()
        return JsonResponse({"success": True, "message": "Dossier supprimé avec succès."})
    except Exception as e:
        traceback.print_exc()
        return _json_error(f"Erreur : {str(e)}", status=500)
