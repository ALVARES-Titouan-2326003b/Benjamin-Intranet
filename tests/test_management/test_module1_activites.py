import io
import json
from datetime import timedelta, timezone as dt_timezone
from unittest.mock import patch

import pytest
from django.contrib.auth.models import Group, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from management.email_manager import _activity_event_payload
from management.models import (
    Activite,
    CategorieDossierAdministratif,
    ChampPersonnaliseDossier,
    HistoriqueRappelActivite,
    RegleRappelActivite,
    TypeActivite,
    ValeurChampPersonnaliseDossier,
)
from management.tasks import check_and_send_activite_reminders
from technique.models import DocumentTechnique, TechnicalProject


@pytest.fixture
def admin_user(db):
    return User.objects.create_superuser(
        username="admin_module1",
        email="admin_module1@example.com",
        password="testpass123",
    )


@pytest.fixture
def responsable(db):
    return User.objects.create_user(
        username="responsable_module1",
        email="responsable_module1@example.com",
        password="testpass123",
    )


@pytest.fixture
def categorie(db):
    return CategorieDossierAdministratif.objects.create(
        nom=CategorieDossierAdministratif.DEFAULT_NOM,
        is_default=True,
    )


@pytest.fixture
def dossier(db, categorie):
    return TechnicalProject.objects.create(
        reference="ADM-001",
        name="Dossier administratif",
        affaire="Dossier administratif",
        categorie=categorie,
    )


@pytest.fixture
def type_activite(db):
    return TypeActivite.objects.create(type="Relance")


def _post_json(client, url, payload):
    return client.post(
        url,
        data=json.dumps(payload),
        content_type="application/json",
    )


def _xlsx_upload(rows, name="dossiers.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        name,
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _xlsx_upload_sheets(sheets, name="dossiers.xlsx"):
    workbook = Workbook()
    first = True
    for title, rows in sheets:
        sheet = workbook.active if first else workbook.create_sheet(title=title)
        sheet.title = title
        first = False
        for row in rows:
            sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        name,
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.mark.django_db
def test_superadmin_does_not_see_admin_calendar_toggle(client, admin_user):
    client.force_login(admin_user)

    response = client.get("/administratif/")

    assert response.status_code == 200
    assert b'data-calendar-scope="ceo"' not in response.content
    assert b"Calendrier administrateur" not in response.content


@pytest.mark.django_db
def test_pole_administratif_sees_admin_calendar_toggle(client, admin_user):
    pole_admin, _ = Group.objects.get_or_create(name="POLE_ADMINISTRATIF")
    user = User.objects.create_user(
        username="membre_admin",
        email="membre_admin@example.com",
        password="testpass123",
    )
    user.groups.add(pole_admin)
    client.force_login(user)

    response = client.get("/administratif/")

    assert response.status_code == 200
    assert b'data-calendar-scope="ceo"' in response.content
    assert "Calendrier administrateur".encode() in response.content


@pytest.mark.django_db
def test_activity_create_update_delete_with_outlook(client, admin_user, responsable, dossier, type_activite):
    client.force_login(admin_user)
    date_value = (timezone.now() + timedelta(days=8)).replace(second=0, microsecond=0)

    payload = {
        "titre": "Relance notaire",
        "dossier": dossier.reference,
        "type": type_activite.type,
        "date": date_value.isoformat(),
        "responsable": str(responsable.pk),
        "statut": "todo",
        "priorite": "high",
        "duree_minutes": 30,
        "client": "Client Test",
        "contact_externe": "notaire@example.com",
        "commentaire": "Vérifier les pièces manquantes.",
        "sync_outlook": True,
    }

    with patch("management.views.create_outlook_event", return_value={"success": True, "event_id": "evt-1"}):
        response = _post_json(client, "/api/create-activity/", payload)

    assert response.status_code == 200
    data = response.json()
    assert data["success"] is True
    activity = Activite.objects.get(pk=data["activity_id"])
    assert activity.titre == "Relance notaire"
    assert activity.responsable == responsable
    assert activity.priorite == "high"
    assert activity.duree_minutes == 30
    assert data["activity"]["duree_minutes"] == 30
    assert data["activity"]["duree_label"] == "30 min"
    assert activity.outlook_event_id == "evt-1"

    update_payload = {
        **payload,
        "titre": "Relance compromis",
        "statut": "in_progress",
        "priorite": "urgent",
        "duree_minutes": 90,
        "client": "Client Test Modifié",
        "sync_outlook": True,
    }
    with patch("management.views.update_outlook_event", return_value={"success": True, "event_id": "evt-1"}):
        response = _post_json(client, f"/api/update-activity/{activity.pk}/", update_payload)

    assert response.status_code == 200
    activity.refresh_from_db()
    assert activity.titre == "Relance compromis"
    assert activity.statut == "in_progress"
    assert activity.priorite == "urgent"
    assert activity.duree_minutes == 90

    with patch("management.views.delete_outlook_event", return_value={"success": True}):
        response = _post_json(client, "/api/delete-activity/", {"activity_id": activity.pk})

    assert response.status_code == 200
    assert response.json()["deleted_count"] == 1
    assert not Activite.objects.filter(pk=activity.pk).exists()


@pytest.mark.django_db
def test_activity_rejects_invalid_duration(client, admin_user, dossier, type_activite):
    client.force_login(admin_user)
    date_value = (timezone.now() + timedelta(days=3)).replace(second=0, microsecond=0)

    response = _post_json(
        client,
        "/api/create-activity/",
        {
            "titre": "Créneau invalide",
            "dossier": dossier.reference,
            "type": type_activite.type,
            "date": date_value.isoformat(),
            "duree_minutes": 45,
        },
    )

    assert response.status_code == 400
    assert "Durée de créneau invalide" in response.json()["message"]


@pytest.mark.django_db
def test_activity_can_be_created_without_dossier(client, admin_user, type_activite):
    client.force_login(admin_user)
    date_value = (timezone.now() + timedelta(days=2)).replace(second=0, microsecond=0)

    response = _post_json(
        client,
        "/api/create-activity/",
        {
            "titre": "Rappel interne",
            "dossier": "",
            "type": type_activite.type,
            "date": date_value.isoformat(),
            "duree_minutes": 60,
            "statut": "todo",
            "priorite": "normal",
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["success"] is True
    activity = Activite.objects.get(pk=data["activity_id"])
    assert activity.dossier is None
    assert activity.responsable == admin_user
    assert data["activity"]["dossier"] == ""


@pytest.mark.django_db
def test_activity_outlook_payload_uses_duration(dossier, type_activite):
    start = timezone.make_aware(timezone.datetime(2026, 7, 1, 9, 30))
    activity = Activite.objects.create(
        id="duration-outlook",
        titre="Rendez-vous notaire",
        dossier=dossier,
        type=type_activite,
        date=start,
        duree_minutes=90,
        date_type="date",
        statut="todo",
        priorite="normal",
    )

    payload = _activity_event_payload(activity)

    assert payload["start"]["dateTime"] == "2026-07-01T09:30:00"
    assert payload["end"]["dateTime"] == "2026-07-01T11:00:00"
    assert "Durée : 1 h 30" in payload["body"]["content"]


@pytest.mark.django_db
def test_calendar_export_ics_uses_activity_duration(client, admin_user, dossier, type_activite, responsable):
    client.force_login(admin_user)
    Activite.objects.create(
        id="ics-export",
        titre="Signature promesse",
        dossier=dossier,
        type=type_activite,
        date=timezone.datetime(2026, 7, 1, 9, 30, tzinfo=dt_timezone.utc),
        duree_minutes=90,
        date_type="date",
        statut="todo",
        priorite="high",
        responsable=admin_user,
        commentaire="Prévoir les pièces annexes.",
        created_by=admin_user,
    )
    Activite.objects.create(
        id="ics-other-month",
        titre="Hors mois",
        dossier=dossier,
        type=type_activite,
        date=timezone.datetime(2026, 8, 1, 9, 30, tzinfo=dt_timezone.utc),
        duree_minutes=60,
        date_type="date",
        statut="todo",
        priorite="normal",
        created_by=admin_user,
    )

    response = client.get("/administratif/calendrier/export.ics?month=7&year=2026")

    assert response.status_code == 200
    assert response["Content-Type"] == "text/calendar; charset=utf-8"
    assert response["Content-Disposition"] == 'attachment; filename="calendrier_administratif.ics"'
    content = response.content.decode("utf-8")
    unfolded_content = content.replace("\r\n ", "")
    assert "BEGIN:VCALENDAR" in content
    assert "BEGIN:VEVENT" in content
    assert "SUMMARY:Signature promesse" in unfolded_content
    assert "DTSTART:20260701T093000Z" in unfolded_content
    assert "DTEND:20260701T110000Z" in unfolded_content
    assert "Créneau : 1 h 30" in unfolded_content
    assert "Responsable : admin_module1" in unfolded_content
    assert "Prévoir les pièces annexes." in unfolded_content
    assert "Hors mois" not in content


@pytest.mark.django_db
def test_calendar_import_ics_creates_activities_and_skips_duplicates(client, admin_user):
    client.force_login(admin_user)
    content = """BEGIN:VCALENDAR
VERSION:2.0
BEGIN:VEVENT
UID:notaire-1@example.com
DTSTART:20260701T093000Z
DTEND:20260701T110000Z
SUMMARY:Signature promesse importée
DESCRIPTION:Prévoir les annexes\\nAppeler le notaire
END:VEVENT
END:VCALENDAR
"""
    uploaded = SimpleUploadedFile("calendrier.ics", content.encode("utf-8"), content_type="text/calendar")

    response = client.post(
        "/administratif/calendrier/import/",
        {
            "file": uploaded,
        },
    )

    assert response.status_code == 302
    activity = Activite.objects.get(titre="Signature promesse importée")
    assert activity.dossier is None
    assert activity.type_id == "Import calendrier"
    assert activity.responsable == admin_user
    assert activity.duree_minutes == 90
    assert "Prévoir les annexes" in activity.commentaire
    assert "UID calendrier : notaire-1@example.com" in activity.commentaire

    uploaded = SimpleUploadedFile("calendrier.ics", content.encode("utf-8"), content_type="text/calendar")
    response = client.post(
        "/administratif/calendrier/import/",
        {
            "file": uploaded,
        },
    )

    assert response.status_code == 302
    assert Activite.objects.filter(titre="Signature promesse importée").count() == 1


@pytest.mark.django_db
def test_calendar_import_ics_can_create_activity_without_dossier(client, admin_user, type_activite):
    client.force_login(admin_user)
    content = """BEGIN:VCALENDAR
VERSION:2.0
BEGIN:VEVENT
UID:external-calendar-1@example.com
DTSTART:20260702T080000Z
DTEND:20260702T090000Z
SUMMARY:Rendez-vous externe importé
DESCRIPTION:Événement sans dossier intranet
END:VEVENT
END:VCALENDAR
"""
    uploaded = SimpleUploadedFile("calendrier-externe.ics", content.encode("utf-8"), content_type="text/calendar")

    response = client.post(
        "/administratif/calendrier/import/",
        {"file": uploaded},
    )

    assert response.status_code == 302
    activity = Activite.objects.get(titre="Rendez-vous externe importé")
    assert activity.dossier is None
    assert activity.type_id == "Import calendrier"
    assert activity.responsable == admin_user
    assert "UID calendrier : external-calendar-1@example.com" in activity.commentaire


@pytest.mark.django_db
def test_admin_dossier_create_update_delete_and_block_when_used(client, admin_user, type_activite, categorie):
    client.force_login(admin_user)

    response = _post_json(
        client,
        "/api/admin-projects/create/",
        {
            "reference": "adm-new",
            "affaire": "Dossier administratif",
            "lot_etage": "Lot 4 / 2e",
            "adresse_bien": "10 rue de la Paix",
            "vendeur": "Vendeur Test",
            "beneficiaire": "Bénéficiaire Test",
            "locataire": "Locataire Test",
            "type_dossier": "vente",
            "activite_metier": "marchand_biens",
            "etat": "promesse",
            "categorie_id": categorie.pk,
            "date_promesse": "2026-07-01",
            "negociation_externe": "Négociation en cours",
            "frais": "1500.25",
            "prix": "125000.50",
            "dg": "5000",
            "date_dg": "2026-07-05",
            "cs_pret": "Condition suspensive active",
            "date_cs_pret": "2026-07-15",
            "date_reiteration": "2026-09-01",
            "acte": "Acte à contrôler",
        },
    )

    assert response.status_code == 200
    data = response.json()
    assert data["success"] is True
    project = TechnicalProject.objects.get(pk=data["project"]["id"])
    assert project.reference == "ADM-NEW"
    assert project.affaire == "Dossier administratif"
    assert project.type_dossier == "vente"
    assert project.activite_metier == "marchand_biens"
    assert project.etat == "promesse"
    assert project.categorie == categorie
    assert str(project.prix) == "125000.50"
    assert str(project.frais) == "1500.25"
    assert str(project.dg) == "5000.00"

    response = _post_json(
        client,
        f"/api/admin-projects/{project.pk}/update/",
        {
            "reference": "ADM-UPD",
            "affaire": "Dossier administratif modifié",
            "type_dossier": "acquisition",
            "activite_metier": "patrimoine",
            "etat": "achete",
            "categorie_id": categorie.pk,
            "prix": "130000",
        },
    )

    assert response.status_code == 200
    project.refresh_from_db()
    assert project.reference == "ADM-UPD"
    assert project.affaire == "Dossier administratif modifié"
    assert project.type_dossier == "acquisition"
    assert project.activite_metier == "patrimoine"
    assert project.etat == "achete"

    Activite.objects.create(
        id="project-blocked",
        titre="Activité liée",
        dossier=project,
        type=type_activite,
        date=timezone.now() + timedelta(days=2),
        date_type="date",
        statut="todo",
        priorite="normal",
        created_by=admin_user,
    )

    response = client.post(f"/api/admin-projects/{project.pk}/delete/")
    assert response.status_code == 409
    assert response.json()["success"] is False
    assert TechnicalProject.objects.filter(pk=project.pk).exists()

    Activite.objects.filter(dossier=project).delete()
    response = client.post(f"/api/admin-projects/{project.pk}/delete/")
    assert response.status_code == 200
    assert response.json()["success"] is True
    assert not TechnicalProject.objects.filter(pk=project.pk).exists()


@pytest.mark.django_db
def test_admin_dossier_accepts_promotion_immobiliere_fields(client, admin_user, categorie):
    client.force_login(admin_user)

    response = _post_json(
        client,
        "/api/admin-projects/create/",
        {
            "reference": "adm-promo",
            "affaire": "Promotion immobilière",
            "type_dossier": "acquisition",
            "activite_metier": "promotion_immobiliere",
            "etat": "promesse",
            "categorie_id": categorie.pk,
            "parcelles": "AB 123, AB 124",
            "premiere_periode": "Janvier 2026",
            "deuxieme_periode": "Février 2026",
            "avenant_1": "Avenant signé",
            "avenant_2": "Avenant à préparer",
            "avenant_3": "Sans objet",
            "depot_permis": "2026-07-01",
            "obtention_permis": "2026-09-15",
            "diags": "Diagnostics en cours",
            "bornage": "Bornage validé",
            "etude_sol_geotechnique": "G2 demandée",
            "etude_pollution": "RAS",
            "etude_impact": "Étude à relire",
            "prorogation": "Prorogation possible",
            "releves_compte": "Relevés transmis",
            "prix": "250000",
        },
    )

    assert response.status_code == 200
    project = TechnicalProject.objects.get(reference="ADM-PROMO")
    assert project.activite_metier == "promotion_immobiliere"
    assert project.get_activite_metier_display() == "Promotion immobilière"
    assert project.categorie == categorie
    assert project.parcelles == "AB 123, AB 124"
    assert project.premiere_periode == "Janvier 2026"
    assert project.deuxieme_periode == "Février 2026"
    assert project.avenant_1 == "Avenant signé"
    assert project.depot_permis.isoformat() == "2026-07-01"
    assert project.obtention_permis.isoformat() == "2026-09-15"
    assert project.diags == "Diagnostics en cours"
    assert project.bornage == "Bornage validé"
    assert project.etude_sol_geotechnique == "G2 demandée"
    assert project.etude_pollution == "RAS"
    assert project.etude_impact == "Étude à relire"
    assert project.prorogation == "Prorogation possible"
    assert project.releves_compte == "Relevés transmis"


@pytest.mark.django_db
def test_admin_dossier_delete_is_blocked_by_technical_links(client, admin_user, categorie):
    client.force_login(admin_user)
    project = TechnicalProject.objects.create(
        reference="ADM-TECH-LINK",
        name="Dossier avec document technique",
        affaire="Dossier avec document technique",
        categorie=categorie,
    )
    DocumentTechnique.objects.create(
        project=project,
        titre="Promesse",
        fichier=SimpleUploadedFile("promesse.txt", b"contenu"),
        created_by=admin_user,
    )

    response = client.post(f"/api/admin-projects/{project.pk}/delete/")

    assert response.status_code == 409
    assert response.json()["success"] is False
    assert "document" in response.json()["message"]
    assert TechnicalProject.objects.filter(pk=project.pk).exists()


@pytest.mark.django_db
def test_admin_overview_dossiers_and_legacy_projects_redirect_are_split(client, admin_user, dossier):
    client.force_login(admin_user)

    with patch("management.views.sync_conversation_journal") as sync_mock:
        overview = client.get("/administratif/")
    dossiers = client.get("/administratif/dossiers/")
    legacy_projects = client.get("/administratif/projets/")
    detail = client.get(f"/administratif/dossiers/{dossier.pk}/")

    assert overview.status_code == 200
    sync_mock.assert_not_called()
    assert dossiers.status_code == 200
    assert legacy_projects.status_code == 302
    assert legacy_projects["Location"] == "/administratif/dossiers/"
    assert detail.status_code == 200
    assert b"admin-projects-data" not in overview.content
    assert b"admin-projects-data" in dossiers.content
    assert dossier.reference.encode() in dossiers.content
    assert dossier.reference.encode() in detail.content


@pytest.mark.django_db
def test_admin_dossiers_export_xlsx(client, admin_user, dossier):
    client.force_login(admin_user)
    TechnicalProject.objects.create(
        reference="ADM-PROMO-EXPORT",
        name="Dossier promotion",
        affaire="Dossier promotion",
        activite_metier="promotion_immobiliere",
        categorie=dossier.categorie,
    )

    response = client.get("/administratif/dossiers/export/")

    assert response.status_code == 200
    assert response["Content-Disposition"] == 'attachment; filename="dossiers_administratifs.xlsx"'
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["Promotion immobilière", "Autres dossiers"]

    promotion_sheet = workbook["Promotion immobilière"]
    other_sheet = workbook["Autres dossiers"]
    assert promotion_sheet["A1"].value == "Référence"
    assert promotion_sheet["B1"].value == "Affaire"
    assert promotion_sheet["A2"].value == "ADM-PROMO-EXPORT"
    assert other_sheet["A2"].value == dossier.reference
    assert other_sheet["B2"].value == dossier.affaire


@pytest.mark.django_db
def test_admin_dossiers_export_pdf(client, admin_user, dossier):
    client.force_login(admin_user)

    response = client.get("/administratif/dossiers/export/pdf/")

    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    assert response["Content-Disposition"] == 'attachment; filename="dossiers_administratifs.pdf"'
    assert response.content.startswith(b"%PDF")


@pytest.mark.django_db
def test_admin_dossiers_import_xlsx_create_and_update(client, admin_user, categorie):
    client.force_login(admin_user)
    uploaded = _xlsx_upload(
        [
            ["Référence", "Affaire", "Type de dossier", "Activité métier", "État", "Catégorie", "Prix", "Date de promesse"],
            ["ADM-IMPORT", "Dossier importé", "Vente", "Marchands de bien", "En cours de promesse", categorie.nom, "125000,50", "01/07/2026"],
        ]
    )

    response = client.post("/administratif/dossiers/import/", {"file": uploaded})

    assert response.status_code == 302
    project = TechnicalProject.objects.get(reference="ADM-IMPORT")
    assert project.affaire == "Dossier importé"
    assert project.type_dossier == "vente"
    assert project.activite_metier == "marchand_biens"
    assert project.etat == "promesse"
    assert project.categorie == categorie
    assert str(project.prix) == "125000.50"
    assert project.date_promesse.isoformat() == "2026-07-01"

    uploaded = _xlsx_upload(
        [
            ["Référence", "Affaire", "Type de dossier", "Activité métier", "État", "Catégorie", "Prix"],
            ["ADM-IMPORT", "Dossier importé modifié", "Acquisition", "Patrimoine", "Acheté", categorie.nom, "130000"],
        ]
    )

    response = client.post("/administratif/dossiers/import/", {"file": uploaded})

    assert response.status_code == 302
    project.refresh_from_db()
    assert project.affaire == "Dossier importé modifié"
    assert project.type_dossier == "acquisition"
    assert project.activite_metier == "patrimoine"
    assert project.etat == "achete"
    assert str(project.prix) == "130000.00"


@pytest.mark.django_db
def test_admin_dossiers_import_xlsx_uses_sheet_to_separate_activity(client, admin_user, categorie):
    client.force_login(admin_user)
    uploaded = _xlsx_upload_sheets(
        [
            (
                "Promotion immobilière",
                [
                    ["Référence", "Affaire", "Type de dossier", "État", "Catégorie"],
                    ["ADM-PROMO-SHEET", "Dossier promotion", "Acquisition", "En cours de promesse", categorie.nom],
                ],
            ),
            (
                "Autres dossiers",
                [
                    ["Référence", "Affaire", "Type de dossier", "État", "Catégorie"],
                    ["ADM-OTHER-SHEET", "Dossier autre", "Vente", "En cours de promesse", categorie.nom],
                ],
            ),
        ]
    )

    response = client.post("/administratif/dossiers/import/", {"file": uploaded})

    assert response.status_code == 302
    assert TechnicalProject.objects.get(reference="ADM-PROMO-SHEET").activite_metier == "promotion_immobiliere"
    assert TechnicalProject.objects.get(reference="ADM-OTHER-SHEET").activite_metier == "marchand_biens"


@pytest.mark.django_db
def test_admin_dossiers_import_rejects_csv(client, admin_user):
    client.force_login(admin_user)
    uploaded = SimpleUploadedFile(
        "dossiers.csv",
        b"Reference,Affaire\nADM-CSV,Dossier CSV\n",
        content_type="text/csv",
    )

    response = client.post("/administratif/dossiers/import/", {"file": uploaded})

    assert response.status_code == 302
    assert not TechnicalProject.objects.filter(reference="ADM-CSV").exists()


@pytest.mark.django_db
def test_admin_dossier_rejects_invalid_choices_and_negative_amounts(client, admin_user, categorie):
    client.force_login(admin_user)
    base_payload = {
        "reference": "ADM-INVALID",
        "affaire": "Dossier invalide",
        "type_dossier": "vente",
        "activite_metier": "marchand_biens",
        "etat": "promesse",
        "categorie_id": categorie.pk,
        "prix": "100000",
    }

    response = _post_json(client, "/api/admin-projects/create/", {**base_payload, "type_dossier": "location"})
    assert response.status_code == 400
    assert "Type de dossier invalide" in response.json()["message"]

    response = _post_json(client, "/api/admin-projects/create/", {**base_payload, "reference": "ADM-NEG", "prix": "-1"})
    assert response.status_code == 400
    assert "Prix ne peut pas être négatif" in response.json()["message"]

    response = _post_json(
        client,
        "/api/admin-projects/create/",
        {**base_payload, "reference": "ADM-DATE", "date_promesse": "01/07/2026"},
    )
    assert response.status_code == 400
    assert "Date de promesse doit être une date valide" in response.json()["message"]


@pytest.mark.django_db
def test_custom_admin_field_definition_and_project_values_are_preserved(client, admin_user, categorie):
    client.force_login(admin_user)

    response = _post_json(
        client,
        "/api/admin-custom-fields/create/",
        {
            "label": "Notaire référent",
            "field_type": "text",
            "show_in_detail": True,
            "show_in_table": True,
            "is_active": True,
            "sort_order": 1,
        },
    )

    assert response.status_code == 200
    field = ChampPersonnaliseDossier.objects.get(label="Notaire référent")

    response = _post_json(
        client,
        "/api/admin-projects/create/",
        {
            "reference": "ADM-CUSTOM",
            "affaire": "Dossier avec champ personnalisé",
            "type_dossier": "vente",
            "activite_metier": "marchand_biens",
            "etat": "promesse",
            "categorie_id": categorie.pk,
            "prix": "100000",
            "custom_fields": {str(field.pk): "Me Martin"},
        },
    )

    assert response.status_code == 200
    project = TechnicalProject.objects.get(reference="ADM-CUSTOM")
    assert ValeurChampPersonnaliseDossier.objects.get(dossier=project, field=field).value == "Me Martin"
    assert response.json()["dossier"]["custom_fields"][str(field.pk)] == "Me Martin"

    response = _post_json(
        client,
        f"/api/admin-custom-fields/{field.pk}/update/",
        {
            "label": "Notaire référent",
            "field_type": "text",
            "show_in_detail": True,
            "show_in_table": True,
            "is_active": False,
            "sort_order": 1,
        },
    )

    assert response.status_code == 200
    field.refresh_from_db()
    assert field.is_active is False
    assert ValeurChampPersonnaliseDossier.objects.get(dossier=project, field=field).value == "Me Martin"


@pytest.mark.django_db
def test_custom_admin_field_validates_choice_values(client, admin_user, categorie):
    client.force_login(admin_user)
    field = ChampPersonnaliseDossier.objects.create(
        label="Risque dossier",
        field_type="choice",
        choices="Faible\nMoyen\nFort",
        show_in_detail=True,
        show_in_table=True,
    )

    response = _post_json(
        client,
        "/api/admin-projects/create/",
        {
            "reference": "ADM-CUSTOM-CHOICE",
            "affaire": "Dossier choix invalide",
            "type_dossier": "vente",
            "activite_metier": "marchand_biens",
            "etat": "promesse",
            "categorie_id": categorie.pk,
            "prix": "100000",
            "custom_fields": {str(field.pk): "Très fort"},
        },
    )

    assert response.status_code == 400
    assert "Risque dossier doit correspondre à un choix autorisé" in response.json()["message"]


@pytest.mark.django_db
def test_custom_admin_field_is_scoped_by_activity(client, admin_user, categorie):
    client.force_login(admin_user)

    response = _post_json(
        client,
        "/api/admin-custom-fields/create/",
        {
            "label": "Dépôt permis personnalisé",
            "activite_metier": "promotion_immobiliere",
            "field_type": "date",
            "show_in_detail": True,
            "show_in_table": True,
            "is_active": True,
        },
    )

    assert response.status_code == 200
    field = ChampPersonnaliseDossier.objects.get(label="Dépôt permis personnalisé")
    assert field.activite_metier == "promotion_immobiliere"
    assert response.json()["field"]["activite_metier_label"] == "Promotion immobilière"

    marchand_response = _post_json(
        client,
        "/api/admin-projects/create/",
        {
            "reference": "ADM-CUSTOM-MARCHAND",
            "affaire": "Dossier marchand",
            "type_dossier": "vente",
            "activite_metier": "marchand_biens",
            "etat": "promesse",
            "categorie_id": categorie.pk,
            "prix": "100000",
            "custom_fields": {str(field.pk): "2026-07-01"},
        },
    )
    promotion_response = _post_json(
        client,
        "/api/admin-projects/create/",
        {
            "reference": "ADM-CUSTOM-PROMO",
            "affaire": "Dossier promotion",
            "type_dossier": "acquisition",
            "activite_metier": "promotion_immobiliere",
            "etat": "promesse",
            "categorie_id": categorie.pk,
            "prix": "200000",
            "custom_fields": {str(field.pk): "2026-07-01"},
        },
    )

    assert marchand_response.status_code == 200
    assert promotion_response.status_code == 200
    marchand_project = TechnicalProject.objects.get(reference="ADM-CUSTOM-MARCHAND")
    promotion_project = TechnicalProject.objects.get(reference="ADM-CUSTOM-PROMO")
    assert not ValeurChampPersonnaliseDossier.objects.filter(dossier=marchand_project, field=field).exists()
    assert ValeurChampPersonnaliseDossier.objects.get(dossier=promotion_project, field=field).value == "2026-07-01"


@pytest.mark.django_db
def test_custom_admin_field_export_and_import_xlsx(client, admin_user, categorie):
    client.force_login(admin_user)
    field = ChampPersonnaliseDossier.objects.create(
        label="Notaire référent",
        field_type="text",
        show_in_table=True,
    )
    project = TechnicalProject.objects.create(
        reference="ADM-CUSTOM-EXPORT",
        name="Dossier export personnalisé",
        affaire="Dossier export personnalisé",
        categorie=categorie,
    )
    ValeurChampPersonnaliseDossier.objects.create(
        dossier=project,
        field=field,
        value="Me Durand",
    )

    response = client.get("/administratif/dossiers/export/")
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook["Autres dossiers"]

    headers = [cell.value for cell in sheet[1]]
    assert "Notaire référent" in headers
    custom_column = headers.index("Notaire référent") + 1
    assert sheet.cell(row=2, column=custom_column).value == "Me Durand"

    uploaded = _xlsx_upload(
        [
            ["Référence", "Affaire", "Type de dossier", "État", "Catégorie", "Notaire référent"],
            ["ADM-CUSTOM-IMPORT", "Dossier import personnalisé", "Vente", "En cours de promesse", categorie.nom, "Me Petit"],
        ]
    )
    response = client.post("/administratif/dossiers/import/", {"file": uploaded})

    assert response.status_code == 302
    imported_project = TechnicalProject.objects.get(reference="ADM-CUSTOM-IMPORT")
    assert ValeurChampPersonnaliseDossier.objects.get(dossier=imported_project, field=field).value == "Me Petit"


@pytest.mark.django_db
def test_gmail_journal_sync_endpoint_runs_on_demand(client, admin_user):
    client.force_login(admin_user)

    with patch(
        "management.views.sync_conversation_journal",
        return_value={"synced": 3, "replied": 1},
    ) as sync_mock:
        response = client.post("/api/gmail-journal/sync/")

    assert response.status_code == 200
    assert response.json() == {"success": True, "synced": 3, "replied": 1}
    sync_mock.assert_called_once_with(admin_user, limit=100)


@pytest.mark.django_db
def test_activity_reminder_is_sent_once(
    settings,
    responsable,
    dossier,
    type_activite,
):
    settings.EMAIL_HOST_USER = "fallback@example.com"
    RegleRappelActivite.objects.get_or_create(timing="before", days=7, defaults={"is_active": True})
    activity = Activite.objects.create(
        id="1001",
        titre="Échéance J-7",
        dossier=dossier,
        type=type_activite,
        date=timezone.now() + timedelta(days=7),
        date_type="date",
        statut="todo",
        priorite="high",
        responsable=responsable,
        created_by=responsable,
    )

    with patch("management.tasks.EmailMessage.send", return_value=1) as send_mock:
        first = check_and_send_activite_reminders()
        second = check_and_send_activite_reminders()

    assert first["rappels_envoyes"] == 1
    assert second["rappels_envoyes"] == 0
    assert second["doublons_ignores"] == 1
    assert send_mock.call_count == 1
    assert HistoriqueRappelActivite.objects.filter(
        activite=activity,
        canal="email",
        jours_avant_echeance=7,
        statut="sent",
    ).count() == 1


@pytest.mark.django_db
def test_finished_activity_does_not_generate_reminder(settings, responsable, dossier, type_activite):
    settings.EMAIL_HOST_USER = "fallback@example.com"
    Activite.objects.create(
        id="1002",
        titre="Activité terminée",
        dossier=dossier,
        type=type_activite,
        date=timezone.now() + timedelta(days=7),
        date_type="date",
        statut="done",
        priorite="normal",
        responsable=responsable,
        created_by=responsable,
    )

    with patch("management.tasks.EmailMessage.send", return_value=1) as send_mock:
        result = check_and_send_activite_reminders()

    assert result["activites_traitees"] == 0
    assert result["rappels_envoyes"] == 0
    assert send_mock.call_count == 0
    assert HistoriqueRappelActivite.objects.count() == 0


@pytest.mark.django_db
def test_activity_reminder_after_due_date_uses_configured_rule(settings, responsable, dossier, type_activite):
    settings.EMAIL_HOST_USER = "fallback@example.com"
    RegleRappelActivite.objects.create(timing="after", days=1, is_active=True)
    activity = Activite.objects.create(
        id="1003",
        titre="Échéance dépassée",
        dossier=dossier,
        type=type_activite,
        date=timezone.now() - timedelta(days=1),
        date_type="date",
        statut="todo",
        priorite="high",
        responsable=responsable,
        created_by=responsable,
    )

    with patch("management.tasks.EmailMessage.send", return_value=1) as send_mock:
        result = check_and_send_activite_reminders()

    assert result["rappels_envoyes"] == 1
    assert send_mock.call_count == 1
    assert HistoriqueRappelActivite.objects.filter(
        activite=activity,
        canal="email",
        jours_avant_echeance=-1,
        statut="sent",
    ).count() == 1


@pytest.mark.django_db
def test_activity_reminder_rule_create_and_delete(client, admin_user):
    client.force_login(admin_user)

    response = client.post("/api/activity-reminder-rules/create/", {"timing": "after", "days": "3"})

    assert response.status_code == 302
    rule = RegleRappelActivite.objects.get(timing="after", days=3)
    assert rule.is_active is True

    response = client.post(f"/api/activity-reminder-rules/{rule.pk}/delete/")

    assert response.status_code == 302
    assert not RegleRappelActivite.objects.filter(pk=rule.pk).exists()
