"""
Tests pour le module invoices.forms
Tests des formulaires de factures et fonctions utilitaires
"""
import pytest
from datetime import datetime
from django.utils import timezone
from django.core.files.uploadedfile import SimpleUploadedFile
from unittest.mock import patch
from django.contrib.auth.models import User, Group
from django.urls import reverse

from invoices.forms import (
    normalize_label,
    best_match,
    build_choices,
    ensure_dossier_exists,
    FactureForm,
    PieceJointeForm,
)
from invoices.models import Facture, Entreprise, Fournisseur, Client, ActeurExterne


class TestNormalizeLabel:
    """Tests pour la fonction normalize_label"""

    def test_normalize_basic_string(self):
        """Teste la normalisation d'une chaîne simple"""
        result = normalize_label("Hello World")
        assert result == "hello world"

    def test_normalize_with_accents(self):
        """Teste la normalisation avec accents"""
        result = normalize_label("Café Français")
        assert result == "cafe francais"

    def test_normalize_with_underscores(self):
        """Teste la normalisation avec underscores"""
        result = normalize_label("Hello_World_Test")
        assert result == "hello world test"

    def test_normalize_with_multiple_spaces(self):
        """Teste la normalisation avec espaces multiples"""
        result = normalize_label("Hello    World")
        assert result == "hello world"

    def test_normalize_empty_string(self):
        """Teste la normalisation d'une chaîne vide"""
        result = normalize_label("")
        assert result == ""

    def test_normalize_none(self):
        """Teste la normalisation de None"""
        result = normalize_label(None)
        assert result is None

    def test_normalize_mixed_case_and_accents(self):
        """Teste avec majuscules, minuscules et accents"""
        result = normalize_label("ComptaBilité_Et Finance")
        assert result == "comptabilite et finance"


class TestBestMatch:
    """Tests pour la fonction best_match"""

    def test_exact_match(self):
        """Teste un match exact"""
        labels = ["Technique", "Administratif", "Comptabilité"]
        result = best_match("Technique", labels)
        assert result == "Technique"

    def test_case_insensitive_match(self):
        """Teste un match insensible à la casse"""
        labels = ["Technique", "Administratif", "Comptabilité"]
        result = best_match("technique", labels)
        assert result == "Technique"

    def test_accent_insensitive_match(self):
        """Teste un match insensible aux accents"""
        labels = ["Comptabilité et Finance"]
        result = best_match("Comptabilite et Finance", labels)
        assert result == "Comptabilité et Finance"

    def test_underscore_to_space_match(self):
        """Teste un match avec underscore → espace"""
        labels = ["Comptabilité et Finance"]
        result = best_match("Comptabilité_et_Finance", labels)
        assert result == "Comptabilité et Finance"

    def test_no_match(self):
        """Teste quand aucun match n'est trouvé"""
        labels = ["Technique", "Administratif"]
        result = best_match("Inexistant", labels)
        assert result is None

    def test_empty_value(self):
        """Teste avec une valeur vide"""
        labels = ["Technique", "Administratif"]
        result = best_match("", labels)
        assert result is None

    def test_empty_labels(self):
        """Teste avec une liste vide de labels"""
        result = best_match("Technique", [])
        assert result is None


class TestBuildChoices:
    """Tests pour la fonction build_choices"""

    def test_build_choices_basic(self):
        """Teste la construction de choices simples"""
        labels = ["Option1", "Option2", "Option3"]
        result = build_choices(labels)
        expected = [("Option1", "Option1"), ("Option2", "Option2"), ("Option3", "Option3")]
        assert result == expected

    def test_build_choices_empty(self):
        """Teste avec une liste vide"""
        result = build_choices([])
        assert result == []

    def test_build_choices_with_accents(self):
        """Teste avec des labels accentués"""
        labels = ["Payée", "En cours", "Archivée"]
        result = build_choices(labels)
        assert result == [("Payée", "Payée"), ("En cours", "En cours"), ("Archivée", "Archivée")]


@pytest.mark.django_db
class TestEnsureDossierExists:
    """Tests pour la fonction ensure_dossier_exists"""

    def test_empty_reference(self):
        """Teste avec une référence vide (ne devrait rien faire)"""
        ensure_dossier_exists("")
        ensure_dossier_exists(None)


@pytest.mark.django_db
class TestFactureForm:
    """Tests pour le formulaire FactureForm"""

    @pytest.fixture
    def mock_enums(self):
        """Mock des ENUMs PostgreSQL"""
        with patch('invoices.forms.get_enum_labels') as mock:
            def side_effect(enum_name):
                if enum_name == "statut":
                    return ["Reçue", "En cours", "Payée", "Archivée"]
                elif enum_name == "poles":
                    return ["Comptabilité et Finance", "Technique", "Administratif"]
                elif enum_name == "type_dossier":
                    return ["Technique", "Administratif"]
                return []
            mock.side_effect = side_effect
            yield mock

    @pytest.fixture
    def valid_form_data(self):
        """Données de formulaire valides"""
        from technique.models import TechnicalProject

        dossier = TechnicalProject.objects.create(reference='DOS-000', name='Dossier Test')

        return {
            'fournisseur_input': 'Fournisseur Test',
            'client_input': 'Client Test',
            'dossier': dossier.id,
            'montant': 1000.0,
            'statut': 'En cours',
            'echeance': '2024-12-31',
            'titre': 'Facture de test',
        }

    def test_form_initialization(self, mock_enums):
        """Teste l'initialisation du formulaire"""
        form = FactureForm()

        assert len(form.fields['statut'].choices) > 0
        assert 'dossier' in form.fields

    def test_form_valid_data(self, mock_enums, valid_form_data):
        """Teste un formulaire avec données valides"""
        form = FactureForm(data=valid_form_data)
        assert form.is_valid()

    def test_form_missing_required_fields(self, mock_enums):
        """Teste avec des champs requis manquants"""
        form = FactureForm(data={})
        assert not form.is_valid()
        assert 'fournisseur_input' in form.errors
        assert 'montant' in form.errors

    @patch('invoices.forms.ensure_dossier_exists')
    def test_form_save_creates_fournisseur(self, mock_ensure, mock_enums, valid_form_data):
        """Teste que save() crée le fournisseur s'il n'existe pas"""
        form = FactureForm(data=valid_form_data)
        assert form.is_valid()

        assert not Fournisseur.objects.filter(id='Fournisseur Test').exists()

        facture = form.save()

        assert Fournisseur.objects.filter(id='Fournisseur Test').exists()
        assert facture.fournisseur == 'Fournisseur Test'

    @patch('invoices.forms.ensure_dossier_exists')
    def test_form_save_creates_client_and_entreprise(self, mock_ensure, mock_enums, valid_form_data):
        """Teste que save() crée le client et l'entreprise"""
        form = FactureForm(data=valid_form_data)
        assert form.is_valid()

        facture = form.save()

        assert Client.objects.filter(id='Client Test').exists()
        assert Entreprise.objects.filter(id='Client Test').exists()
        assert facture.client.id == 'Client Test'

    @patch('invoices.forms.ensure_dossier_exists')
    def test_form_save_empty_client_becomes_divers(self, mock_ensure, mock_enums, valid_form_data):
        """Teste qu'un client vide devient 'DIVERS'"""
        valid_form_data['client_input'] = ''
        form = FactureForm(data=valid_form_data)
        assert form.is_valid()

        facture = form.save()

        assert facture.client.id == 'DIVERS'
        assert Entreprise.objects.filter(id='DIVERS').exists()

    @patch('invoices.forms.ensure_dossier_exists')
    def test_form_save_generates_facture_id(self, mock_ensure, mock_enums, valid_form_data):
        """Teste que save() génère un ID de facture"""
        form = FactureForm(data=valid_form_data)
        assert form.is_valid()

        facture = form.save()

        assert facture.id is not None
        assert facture.id.startswith('FAC-')
        assert len(facture.id) == 12

    @patch('invoices.forms.ensure_dossier_exists')
    def test_form_save_generates_dossier_reference(self, mock_ensure, mock_enums, valid_form_data):
        """Teste que save() génère une référence dossier"""
        form = FactureForm(data=valid_form_data)
        assert form.is_valid()
        facture = form.save()

        assert facture.dossier is not None
        assert facture.dossier.startswith('DOS-')
        mock_ensure.assert_called_once_with(facture.dossier)

    @patch('invoices.forms.ensure_dossier_exists')
    def test_form_save_echeance_with_timezone(self, mock_ensure, mock_enums, valid_form_data):
        """Teste que l'échéance est correctement convertie en datetime aware"""
        form = FactureForm(data=valid_form_data)
        assert form.is_valid()

        facture = form.save()


        assert isinstance(facture.echeance, datetime)

        assert timezone.is_aware(facture.echeance)
        assert facture.echeance.hour == 12
        assert facture.echeance.minute == 0

    @patch('invoices.forms.ensure_dossier_exists')
    def test_form_edit_existing_facture(self, mock_ensure, mock_enums, valid_form_data):
        """Teste l'édition d'une facture existante"""


        acteur = ActeurExterne.objects.create(id='OLD-CLIENT')
        client = Client.objects.create(id=acteur)
        Entreprise.objects.create(id=client, nom='Old Client')

        acteur_f = ActeurExterne.objects.create(id='OLD-FOURNISSEUR')
        fournisseur = Fournisseur.objects.create(id=acteur_f, nom='Old Fournisseur')

        from technique.models import TechnicalProject
        old_dossier = TechnicalProject.objects.create(reference='DOS-OLD', name='Dossier Old')

        facture = Facture.objects.create(
            id='FAC-EXISTING',
            fournisseur=fournisseur,
            client=client,
            montant=500,
            statut='ongoing',
            dossier=old_dossier,
        )

        valid_form_data['fournisseur_input'] = 'New Fournisseur'
        valid_form_data['montant'] = 2000

        form = FactureForm(data=valid_form_data, instance=facture)
        assert form.is_valid()

        updated = form.save()


        assert updated.id == 'FAC-EXISTING'
        assert updated.fournisseur == 'New Fournisseur'
        assert updated.montant == 2000

    @patch('invoices.forms.ensure_dossier_exists')
    def test_form_collaborateur_assignment(self, mock_ensure, mock_enums, valid_form_data):
        """Teste l'assignation d'un collaborateur"""
        from django.contrib.auth.models import Group

        collab_group = Group.objects.create(name='COLLABORATEUR')
        user = User.objects.create_user(username='collab1', email='collab@test.com')
        user.groups.add(collab_group)

        valid_form_data['collaborateur'] = user.id

        form = FactureForm(data=valid_form_data)
        assert form.is_valid()

        facture = form.save()

        assert facture.collaborateur == user




@pytest.mark.django_db
class TestPieceJointeForm:
    """Tests pour le formulaire PieceJointeForm"""

    @pytest.fixture
    def valid_pdf_file(self):
        """Fichier PDF valide pour les tests"""
        return SimpleUploadedFile(
            "test.pdf",
            b"%PDF-1.4 fake pdf content",
            content_type="application/pdf"
        )

    @pytest.fixture
    def facture(self):
        """Facture de test"""
        from technique.models import TechnicalProject

        acteur = ActeurExterne.objects.create(id='TEST-CLIENT')
        client = Client.objects.create(id=acteur)
        Entreprise.objects.create(id=client, nom='Test Client')

        acteur_f = ActeurExterne.objects.create(id='TEST-FOURNISSEUR')
        fournisseur = Fournisseur.objects.create(id=acteur_f, nom='Test Fournisseur')

        dossier = TechnicalProject.objects.create(reference='DOS-TEST', name='Dossier Test')

        return Facture.objects.create(
            id='FAC-TEST',
            fournisseur=fournisseur,
            client=client,
            montant=1000,
            statut='ongoing',
            dossier=dossier,
        )

    def test_form_valid_pdf(self, facture, valid_pdf_file):
        """Teste avec un fichier PDF valide"""
        form = PieceJointeForm(
            data={},
            files={'fichier': valid_pdf_file}
        )
        assert form.is_valid()

    def test_form_invalid_file_type(self, facture):
        """Teste avec un type de fichier invalide"""
        invalid_file = SimpleUploadedFile(
            "test.txt",
            b"This is not a PDF",
            content_type="text/plain"
        )

        form = PieceJointeForm(
            data={},
            files={'fichier': invalid_file}
        )

        assert not form.is_valid()
        assert 'fichier' in form.errors
        assert 'PDF' in str(form.errors['fichier'])

    def test_form_file_too_large(self, facture):
        """Teste avec un fichier trop gros (>10MB)"""

        large_content = b"x" * (11 * 1024 * 1024)
        large_file = SimpleUploadedFile(
            "large.pdf",
            large_content,
            content_type="application/pdf"
        )

        form = PieceJointeForm(
            data={},
            files={'fichier': large_file}
        )

        assert not form.is_valid()
        assert 'fichier' in form.errors
        assert '10 Mo' in str(form.errors['fichier'])

    def test_form_no_file_is_valid(self, facture):
        """Teste que le formulaire est valide sans fichier (optionnel)"""
        form = PieceJointeForm(data={})
        assert form.is_valid()

    def test_form_pdf_extension_validation(self, facture):
        """Teste la validation via l'extension .pdf"""

        pdf_file = SimpleUploadedFile(
            "document.pdf",
            b"fake content",
            content_type="application/octet-stream"
        )

        form = PieceJointeForm(
            data={},
            files={'fichier': pdf_file}
        )

        assert form.is_valid()


@pytest.mark.django_db
class TestFactureListViewFiltersAndExport:
    """Tests pour liste de factures : filtre client/dossier + export Excel/PDF"""

    @pytest.fixture(autouse=True)
    def setup_user(self):
        self.user = User.objects.create_user(username='finance', password='finance', is_staff=True)

    def _create_test_invoice(self, reference='PRJ-001', project_name='Projet Exemple', client_id='CLIENT-1', client_name='Client Exemple'):
        from technique.models import TechnicalProject

        dossier = TechnicalProject.objects.create(reference=reference, name=project_name)
        acteur = ActeurExterne.objects.create(id=client_id)
        client = Client.objects.create(id=acteur)
        Entreprise.objects.create(id=client, nom=client_name)

        acte_fournisseur = ActeurExterne.objects.create(id='FOURNISSEUR-1')
        fournisseur = Fournisseur.objects.create(id=acte_fournisseur, nom='Fournisseur Test')

        facture = Facture.objects.create(
            id='FAC-TEST',
            fournisseur=fournisseur,
            client=client,
            montant=2500.0,
            statut='ongoing',
            dossier=dossier,
        )
        return facture

    def test_client_filter(self, client):
        self._create_test_invoice(client_name='NomClientX')
        client.force_login(self.user)

        response = client.get(reverse('invoices:list'), {'client': 'NomClientX'})
        assert response.status_code == 200
        assert 'FAC-TEST' in response.content.decode('utf-8')

        response = client.get(reverse('invoices:list'), {'client': 'CLIENT-1'})
        assert response.status_code == 200
        assert 'FAC-TEST' in response.content.decode('utf-8')

    def test_dossier_filter(self, client):
        self._create_test_invoice(reference='DOS-AAA', project_name='Mon Dossier')
        client.force_login(self.user)

        response = client.get(reverse('invoices:list'), {'dossier': 'DOS-AAA'})
        assert response.status_code == 200
        assert 'FAC-TEST' in response.content.decode('utf-8')

        response = client.get(reverse('invoices:list'), {'dossier': 'Mon Dossier'})
        assert response.status_code == 200
        assert 'FAC-TEST' in response.content.decode('utf-8')

    def test_export_xlsx_pdf(self, client):
        self._create_test_invoice()
        client.force_login(self.user)

        response_xlsx = client.get(reverse('invoices:list'), {'export': 'xlsx'})
        assert response_xlsx.status_code == 200
        assert response_xlsx['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        from openpyxl import load_workbook
        import io as _io

        f = _io.BytesIO(response_xlsx.content)
        wb = load_workbook(f)
        assert 'Factures' in wb.sheetnames
        ws = wb['Factures']
        assert ws['A1'].value == 'ID'

        response_pdf = client.get(reverse('invoices:list'), {'export': 'pdf'})
        assert response_pdf.status_code == 200
        assert response_pdf['Content-Type'] == 'application/pdf'
        assert response_pdf.content[:4] == b'%PDF'


@pytest.mark.django_db
class TestCollaborateurInvoicePermissions:
    @pytest.fixture(autouse=True)
    def setup(self):
        from technique.models import TechnicalProject

        self.collab_group = Group.objects.get_or_create(name='COLLABORATEUR')[0]
        self.finance_user = User.objects.create_user(username='finance2', password='finance2', is_staff=True)
        self.collaborateur = User.objects.create_user(username='collab', password='collab', email='collab@example.com')
        self.other_collaborateur = User.objects.create_user(username='other-collab', password='collab', email='other@example.com')
        self.collaborateur.groups.add(self.collab_group)
        self.other_collaborateur.groups.add(self.collab_group)
        self.dossier = TechnicalProject.objects.create(reference='DOS-COLLAB', name='Dossier Collaborateur')

    def _create_invoice(self, invoice_id, collaborateur=None, created_by=None, status='ongoing'):
        acteur_client = ActeurExterne.objects.create(id=f'CLIENT-{invoice_id}')
        client = Client.objects.create(id=acteur_client)
        Entreprise.objects.create(id=client, nom=f'Client {invoice_id}')

        acteur_fournisseur = ActeurExterne.objects.create(id=f'FOURN-{invoice_id}')
        fournisseur = Fournisseur.objects.create(id=acteur_fournisseur, nom=f'Fournisseur {invoice_id}')

        return Facture.objects.create(
            id=invoice_id,
            fournisseur=fournisseur,
            client=client,
            montant=100.0,
            statut=status,
            dossier=self.dossier,
            collaborateur=collaborateur,
            created_by=created_by,
            titre=f'Titre {invoice_id}',
        )

    def test_collaborateur_list_shows_only_assigned_invoices_and_create_button(self, client):
        own_invoice = self._create_invoice('FAC-OWN', collaborateur=self.collaborateur, created_by=self.other_collaborateur)
        self._create_invoice('FAC-OTHER', collaborateur=self.other_collaborateur, created_by=self.collaborateur)

        client.force_login(self.collaborateur)
        response = client.get(reverse('invoices:list'))

        content = response.content.decode('utf-8')
        assert response.status_code == 200
        assert own_invoice.id in content
        assert 'FAC-OTHER' not in content
        assert reverse('invoices:create') in content
        assert 'Export Excel' not in content
        assert 'Lancer la relance manuelle' not in content
        assert 'name="invoice_ids"' not in content

    def test_collaborateur_can_view_only_assigned_invoice_detail_and_sees_edit_button(self, client):
        own_invoice = self._create_invoice('FAC-DETAIL-OWN', collaborateur=self.collaborateur)
        other_invoice = self._create_invoice('FAC-DETAIL-OTHER', collaborateur=self.other_collaborateur)

        client.force_login(self.collaborateur)

        own_response = client.get(reverse('invoices:detail', args=[own_invoice.pk]))
        own_content = own_response.content.decode('utf-8')
        assert own_response.status_code == 200
        assert reverse('invoices:edit', args=[own_invoice.pk]) in own_content

        other_response = client.get(reverse('invoices:detail', args=[other_invoice.pk]))
        assert other_response.status_code == 404

    @patch('invoices.forms.ensure_dossier_exists')
    def test_collaborateur_can_create_invoice_but_cannot_set_status_or_assignment(self, mock_ensure, client):
        client.force_login(self.collaborateur)

        response = client.post(reverse('invoices:create'), data={
            'fournisseur_input': 'Fournisseur Creation',
            'client_input': 'Client Creation',
            'dossier': self.dossier.pk,
            'montant': '123.45',
            'echeance': '2026-05-01',
            'titre': 'Facture collaborative',
            'statut': 'paid',
            'collaborateur': self.other_collaborateur.pk,
        })

        facture = Facture.objects.get(titre='Facture collaborative')
        assert response.status_code == 302
        assert facture.created_by == self.collaborateur
        assert facture.collaborateur == self.collaborateur
        assert facture.statut == 'ongoing'

    def test_collaborateur_edit_form_hides_restricted_fields(self, client):
        invoice = self._create_invoice('FAC-FORM', collaborateur=self.collaborateur)

        client.force_login(self.collaborateur)
        response = client.get(reverse('invoices:edit', args=[invoice.pk]))

        content = response.content.decode('utf-8')
        assert response.status_code == 200
        assert 'name="collaborateur"' in content
        assert 'disabled' in content
        assert 'name="statut"' not in content

    def test_collaborateur_can_edit_assigned_invoice_but_not_other_invoice(self, client):
        own_invoice = self._create_invoice('FAC-EDIT-OWN', collaborateur=self.collaborateur, status='ongoing')
        other_invoice = self._create_invoice('FAC-EDIT-OTHER', collaborateur=self.other_collaborateur, status='ongoing')

        client.force_login(self.collaborateur)

        response = client.post(reverse('invoices:edit', args=[own_invoice.pk]), data={
            'fournisseur_input': own_invoice.fournisseur_id,
            'client_input': own_invoice.client_id,
            'dossier': self.dossier.pk,
            'montant': '999.99',
            'echeance': '2026-06-01',
            'titre': 'Titre modifie',
            'statut': 'paid',
            'collaborateur': self.other_collaborateur.pk,
        })

        own_invoice.refresh_from_db()
        assert response.status_code == 302
        assert own_invoice.montant == 999.99
        assert own_invoice.titre == 'Titre modifie'
        assert own_invoice.statut == 'ongoing'
        assert own_invoice.collaborateur == self.collaborateur

        forbidden_response = client.post(reverse('invoices:edit', args=[other_invoice.pk]), data={
            'fournisseur_input': other_invoice.fournisseur_id,
            'client_input': other_invoice.client_id,
            'dossier': self.dossier.pk,
            'montant': '500.00',
            'echeance': '2026-06-01',
            'titre': 'Interdit',
        })
        assert forbidden_response.status_code == 404
